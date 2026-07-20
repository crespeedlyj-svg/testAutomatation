"""
PUR(구매관리) 전체 테스트 시나리오 생성 → 시나리오_1차.xlsx
"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
import os

# ─── 색상 정의 ─────────────────────────────────────────────────────────────
HEADER_BG     = "1A3C72"  # 헤더 배경 (진파랑)
HEADER_FONT   = "FFFFFF"  # 헤더 글자 (흰색)
INTEG_BG      = "E8F0FE"  # 통합 행 배경
UNIT_BG       = "FFF8E1"  # 단위 행 배경
PASS_BG       = "E6F4EA"
FAIL_BG       = "FCE8E6"
SEP_BG        = "2059A3"  # 화면 구분자 배경 (파랑)
SEP_FONT      = "FFFFFF"

thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ─── 컬럼 정의 ────────────────────────────────────────────────────────────
HEADERS = [
    "순번", "구분", "시나리오ID", "화면ID", "화면명",
    "중분류", "소분류", "메뉴명", "CRUD", "API URL",
    "시나리오명", "시나리오흐름", "입력값", "예상결과",
    "테스트결과", "실패사유", "비고"
]
COL_WIDTHS = [
     6, 8, 24, 14, 20,
    14, 16, 22, 10, 40,
    28, 48, 36, 36,
    10, 24, 16
]

# ─── PUR 화면 정의 ─────────────────────────────────────────────────────────
# (화면ID, 화면명, 중분류, 소분류, 메뉴명, 주요URL, crud목록)
# crud목록 = [(crud, url_suffix, 시나리오명_suffix)]
PUR_SCREENS = [
    # ── 구매 본 업무 ──────────────────────────────────────────────────
    ("pur_0010M", "부대비용관리",    "구매관리", "부대비용",   "부대비용관리",
     "/mis/pur/pur0010",
     [("SELECT","getLista","부대비용 목록 조회"),
      ("INSERT","setData","신규 부대비용 등록"),
      ("UPDATE","setData","부대비용 수정"),
      ("DELETE","setData","부대비용 삭제")]),

    ("pur_0110M", "구매요구현황",    "구매관리", "구매요구",   "구매요구현황",
     "/mis/pur/pur0110",
     [("SELECT","getList","구매요구 목록 조회"),
      ("SELECT","getList","조건별 필터 조회")]),

    ("pur_0111M", "구매요구(내자)",  "구매관리", "구매요구",   "구매요구(내자)",
     "/mis/pur/pur0111",
     [("SELECT","getData","구매요구 상세 조회"),
      ("INSERT","setData","신규 구매요구 등록"),
      ("UPDATE","setData","구매요구 수정"),
      ("DELETE","delData","구매요구 삭제"),
      ("UPDATE","uptChk","구매요구 확인 처리")]),

    ("pur_0112M", "구매요구(일반)",  "구매관리", "구매요구",   "구매요구(일반)",
     "/mis/pur/pur0112",
     [("SELECT","getData","구매요구 상세 조회"),
      ("INSERT","setData","신규 구매요구 등록"),
      ("UPDATE","setData","구매요구 수정"),
      ("DELETE","delData","구매요구 삭제")]),

    ("pur_0121M", "구매요구(외자)",  "구매관리", "구매요구",   "구매요구(외자)",
     "/mis/pur/pur0121",
     [("SELECT","getData","외자 구매요구 조회"),
      ("INSERT","setData","외자 구매요구 등록"),
      ("UPDATE","setData","외자 구매요구 수정"),
      ("DELETE","delData","외자 구매요구 삭제")]),

    ("pur_0140M", "계속계약요구현황","구매관리", "계속계약",   "계속계약요구현황",
     "/mis/pur/pur0140",
     [("SELECT","getList","계속계약 목록 조회"),
      ("SELECT","getList","기간 조건 필터 조회")]),

    ("pur_0141M", "계속계약요구신청","구매관리", "계속계약",   "계속계약요구신청",
     "/mis/pur/pur0141",
     [("SELECT","getData","계속계약 상세 조회"),
      ("INSERT","setData","계속계약요구 신규 등록"),
      ("UPDATE","setData","계속계약요구 수정"),
      ("DELETE","delData","계속계약요구 삭제")]),

    ("pur_0150M", "계약변경(해지)현황","구매관리","계약변경",  "계약변경(해지)현황",
     "/mis/pur/pur0150",
     [("SELECT","getList","계약변경 목록 조회")]),

    ("pur_0151M", "계약변경(해지)등록","구매관리","계약변경",  "계약변경(해지)등록",
     "/mis/pur/pur0151",
     [("SELECT","getData","계약변경 상세 조회"),
      ("INSERT","setData","계약변경 신규 등록"),
      ("UPDATE","setData","계약변경 수정"),
      ("DELETE","delData","계약변경 삭제"),
      ("UPDATE","uptChk","계약변경 확인 처리")]),

    ("pur_0152M", "예산변경 목록",   "구매관리", "예산변경",   "예산변경 목록",
     "/mis/pur/pur0152",
     [("SELECT","getList","예산변경 목록 조회"),
      ("SELECT","getList","기간/상태 조건 조회")]),

    ("pur_0153M", "예산변경 상세",   "구매관리", "예산변경",   "예산변경 상세",
     "/mis/pur/pur0153",
     [("SELECT","getData","예산변경 상세 조회"),
      ("SELECT","getDetailData","예산변경 세부 내역 조회"),
      ("INSERT","setData","예산변경 신규 등록"),
      ("UPDATE","setData","예산변경 수정"),
      ("DELETE","delData","예산변경 삭제"),
      ("SELECT","checkBudgBaln","예산잔액 확인")]),

    ("pur_0210M", "구매진행현황",    "구매관리", "구매진행",   "구매진행현황",
     "/mis/pur/pur0210",
     [("SELECT","getPurData","구매진행현황 조회"),
      ("SELECT","getPurData","조건별 진행현황 조회")]),

    ("pur_0211M", "구매품의",        "구매관리", "구매품의",   "구매품의",
     "/mis/pur/pur0211",
     [("SELECT","getRgstList","구매품의 목록 조회"),
      ("SELECT","getRgstInfo","구매품의 상세 조회"),
      ("INSERT","saveAdd","구매품의 신규 등록"),
      ("UPDATE","saveUpt","구매품의 수정"),
      ("DELETE","saveDel","구매품의 삭제"),
      ("UPDATE","setCancel","구매품의 취소")]),

    ("pur_0300M", "구매계약현황",    "구매관리", "구매계약",   "구매계약현황",
     "/mis/pur/pur0300",
     [("SELECT","getList","구매계약 목록 조회"),
      ("SELECT","getList","계약상태별 조회")]),

    ("pur_0301M", "RCMS계약현황",    "구매관리", "구매계약",   "RCMS계약현황",
     "/mis/pur/pur0300",
     [("SELECT","getList","RCMS 계약 목록 조회"),
      ("SELECT","getReslEnblYn","낙찰가능여부 확인")]),

    ("pur_0311M", "계약등록",        "구매관리", "구매계약",   "계약등록",
     "/mis/pur/pur0311",
     [("SELECT","getData","계약 상세 조회"),
      ("INSERT","setData","계약 신규 등록"),
      ("INSERT","setMst","계약 마스터 등록"),
      ("UPDATE","setDtl","계약 세부내역 수정"),
      ("DELETE","delData","계약 삭제"),
      ("UPDATE","purEndData","계약 종료 처리"),
      ("UPDATE","uptChk","계약 확인 처리")]),

    ("pur_0321M", "구매계약(외자)",  "구매관리", "구매계약",   "구매계약(외자)",
     "/mis/pur/pur0321",
     [("SELECT","getData","외자계약 조회"),
      ("INSERT","setData","외자계약 등록"),
      ("UPDATE","setMst","외자계약 마스터 수정"),
      ("DELETE","delData","외자계약 삭제")]),

    ("pur_0350M", "계약구매물품현황","구매관리", "구매계약",   "계약구매물품현황",
     "/mis/pur/pur0350",
     [("SELECT","getList","계약구매물품 목록 조회"),
      ("SELECT","getList","조건별 물품현황 조회")]),

    ("pur_0400M", "검수현황",        "구매관리", "검수",       "검수현황",
     "/mis/pur/pur0400",
     [("SELECT","getList","검수 목록 조회"),
      ("SELECT","getList","기간/상태 조건 조회")]),

    ("pur_0411M", "검수등록",        "구매관리", "검수",       "검수등록",
     "/mis/pur/pur0411",
     [("SELECT","getData","검수 상세 조회"),
      ("INSERT","setExmntData","검수 등록"),
      ("DELETE","delData","검수 삭제"),
      ("UPDATE","setAstCd","자산코드 설정")]),

    ("pur_0430M", "기성/완료 검사조서 현황","구매관리","검수","기성완료검사조서현황",
     "/mis/pur/pur0430",
     [("SELECT","getList","기성검사조서 목록 조회")]),

    ("pur_0431M", "기성/완료 검사조서","구매관리","검수",      "기성완료검사조서",
     "/mis/pur/pur0431",
     [("INSERT","setExmntData","검사조서 등록"),
      ("DELETE","delData","검사조서 삭제"),
      ("UPDATE","uptChk","검사조서 확인 처리")]),

    ("pur_0440M", "납품현황",        "구매관리", "납품",       "납품현황",
     "/mis/pur/pur0440",
     [("SELECT","getListRqst","일반 납품현황 조회"),
      ("SELECT","getListCheap","직접구매 납품현황 조회"),
      ("SELECT","getListMro","MRO 납품현황 조회")]),

    ("pur_0500M", "대금지급신청 목록","구매관리","대금지급",   "대금지급신청목록",
     "/mis/pur/pur0500",
     [("SELECT","getList","대금지급신청 목록 조회"),
      ("SELECT","getList","기간 조건 조회")]),

    ("pur_0521M", "대금지급신청",    "구매관리", "대금지급",   "대금지급신청",
     "/mis/pur/pur0521",
     [("SELECT","getSchParams","검색 파라미터 조회"),
      ("SELECT","getData","대금지급신청 상세 조회"),
      ("INSERT","setData","대금지급신청 등록"),
      ("UPDATE","setChangeCtrl","대금지급 제어정보 수정"),
      ("DELETE","delData","대금지급신청 삭제")]),

    ("pur_0591M", "구매미지급 관리", "구매관리", "대금지급",   "구매미지급관리",
     "/mis/pur/pur0591",
     [("SELECT","getList","미지급 목록 조회"),
      ("UPDATE","setData","미지급 수정"),
      ("DELETE","delData","미지급 삭제")]),

    ("pur_0610M", "입찰공고등록 현황","구매관리","입찰",       "입찰공고등록현황",
     "/mis/pur/pur0610",
     [("SELECT","getList","입찰공고 목록 조회"),
      ("SELECT","getList","입찰상태별 조회")]),

    ("pur_0611M", "입찰공고등록",    "구매관리", "입찰",       "입찰공고등록",
     "/mis/pur/pur0611",
     [("SELECT","getData","입찰공고 상세 조회"),
      ("SELECT","getList","공고 목록 조회"),
      ("INSERT","setData","입찰공고 신규 등록"),
      ("DELETE","delData","입찰공고 삭제"),
      ("UPDATE","uptAnncStat","입찰공고 상태 변경")]),

    ("pur_0710M", "수의시담관리",    "구매관리", "수의시담",   "수의시담관리",
     "/mis/pur/pur0710",
     [("SELECT","getList","수의시담 목록 조회"),
      ("SELECT","getList","기간 조건 조회")]),

    ("pur_0711M", "수의시담등록",    "구매관리", "수의시담",   "수의시담등록",
     "/mis/pur/pur0711",
     [("SELECT","getData","수의시담 상세 조회"),
      ("INSERT","setData","수의시담 신규 등록"),
      ("DELETE","delData","수의시담 삭제"),
      ("UPDATE","setAppro","수의시담 승인"),
      ("UPDATE","setReject","수의시담 반려"),
      ("UPDATE","setAccp","수의시담 접수")]),

    ("pur_0910M", "직접구매요구현황","구매관리", "직접구매",   "직접구매요구현황",
     "/mis/pur/pur0910",
     [("SELECT","getList","직접구매요구 목록 조회"),
      ("SELECT","getList","기간/상태 조건 조회")]),

    ("pur_0911M", "직접구매요구신청","구매관리", "직접구매",   "직접구매요구신청",
     "/mis/pur/pur0911",
     [("SELECT","getSchParams","검색 파라미터 조회"),
      ("SELECT","getCheapData","직접구매 데이터 조회"),
      ("INSERT","setCheapData","직접구매요구 신청 등록"),
      ("INSERT","setExmntData","직접구매 검수 등록"),
      ("DELETE","delData","직접구매요구 삭제"),
      ("UPDATE","uptChk","직접구매요구 확인 처리")]),

    ("pur_0940M", "직접구매검수및대금지급","구매관리","직접구매","직접구매검수대금지급",
     "/mis/pur/pur0940",
     [("SELECT","getList","직접구매 검수 목록 조회"),
      ("UPDATE","uptInspCnrm","검수 확인 처리")]),

    # ── MRO 업무 ──────────────────────────────────────────────────────
    ("pur_1240M", "MRO 검수현황",    "구매관리", "MRO",        "MRO검수현황",
     "/mis/pur/pur_1240_007",
     [("SELECT","getList","MRO 검수 목록 조회"),
      ("SELECT","getList","기간/상태 조건 조회")]),

    ("pur_1250M", "MRO검수",         "구매관리", "MRO",        "MRO검수",
     "/mis/pur/pur1270007",
     [("SELECT","getMroRqstMst","MRO 요청 마스터 조회"),
      ("SELECT","getMroDtlList","MRO 상세 목록 조회"),
      ("INSERT","save","MRO 검수 저장"),
      ("DELETE","delete","MRO 검수 삭제")]),

    ("pur_1260M", "MRO구매취소현황", "구매관리", "MRO",        "MRO구매취소현황",
     "/mis/pur/pur1260007",
     [("SELECT","getList","MRO 구매취소 목록 조회")]),

    ("pur_1270M", "MRO구매취소신청서","구매관리","MRO",        "MRO구매취소신청서",
     "/mis/pur/pur1270007",
     [("SELECT","getMroRqstMst","MRO 취소요청 조회"),
      ("INSERT","save","MRO 취소신청 등록"),
      ("DELETE","delete","MRO 취소신청 삭제")]),

    ("pur_1340M", "MRO발주변경현황", "구매관리", "MRO",        "MRO발주변경현황",
     "/mis/pur/pur1340007",
     [("SELECT","getList","MRO 발주변경 목록 조회")]),

    ("pur_1350M", "MRO발주변경",     "구매관리", "MRO",        "MRO발주변경",
     "/mis/pur/pur1350007",
     [("SELECT","getMroRqstMst","MRO 발주변경 마스터 조회"),
      ("SELECT","getMroDtlList","MRO 발주변경 상세 조회"),
      ("SELECT","getMroModMst","MRO 변경 마스터 조회"),
      ("INSERT","save","MRO 발주변경 등록"),
      ("UPDATE","send","MRO 발주변경 발송"),
      ("DELETE","tmDelete","MRO 발주변경 임시삭제")]),

    ("pur_5110M", "MRO 구매요구현황","구매관리", "MRO",        "MRO구매요구현황",
     "/mis/pur/pur5110",
     [("SELECT","getList","MRO 구매요구 목록 조회"),
      ("SELECT","getList","기간 조건 조회")]),

    ("pur_5115M", "MRO 구매요구신청","구매관리", "MRO",        "MRO구매요구신청",
     "/mis/pur/pur5115",
     [("SELECT","getData","MRO 구매요구 상세 조회"),
      ("INSERT","setData","MRO 구매요구 신청 등록"),
      ("DELETE","delData","MRO 구매요구 삭제"),
      ("SELECT","getDlvrPlcCd","납품장소 코드 조회")]),

    ("pur_5120M", "MRO 요구물품취소","구매관리", "MRO",        "MRO요구물품취소",
     "/mis/pur/pur5120",
     [("SELECT","getList","MRO 요구물품취소 목록 조회"),
      ("UPDATE","setCnclPurRqstData","MRO 요구물품 취소 처리")]),

    ("pur_5130M", "MRO 검수및대금지급 현황","구매관리","MRO",  "MRO검수대금지급현황",
     "/mis/pur/pur5130",
     [("SELECT","getList","MRO 검수대금지급 목록 조회"),
      ("SELECT","getList","기간 조건 조회")]),

    ("pur_5135M", "MRO 검수 및 대금지급신청","구매관리","MRO","MRO검수대금지급신청",
     "/mis/pur/pur5135",
     [("SELECT","getSchParams","검색 파라미터 조회"),
      ("SELECT","getData","MRO 검수대금지급 상세 조회"),
      ("INSERT","setData","MRO 검수대금지급 등록"),
      ("DELETE","delData","MRO 검수대금지급 삭제"),
      ("SELECT","chkSaftyYn","안전 여부 확인")]),

    ("pur_5140M", "MRO 상품계약/변경현황","구매관리","MRO",    "MRO상품계약변경현황",
     "/mis/pur/pur5140",
     [("SELECT","getList","MRO 상품계약 목록 조회")]),

    ("pur_5145M", "MRO 상품계약/변경요구","구매관리","MRO",    "MRO상품계약변경요구",
     "/mis/pur/pur5145",
     [("SELECT","getData","MRO 상품계약 상세 조회"),
      ("INSERT","setData","MRO 상품계약 등록"),
      ("DELETE","delData","MRO 상품계약 삭제")]),

    ("pur_5160M", "MRO 납품확인및검사요청목록","구매관리","MRO","MRO납품확인검사요청",
     "/mis/pur/pur5160",
     [("SELECT","getList","MRO 납품확인 목록 조회"),
      ("UPDATE","setInspCnrm","MRO 검사 확인 처리"),
      ("UPDATE","setPurCancel","MRO 구매 취소 처리")]),

    ("pur_5170M", "MRO발주현황",     "구매관리", "MRO",        "MRO발주현황",
     "/mis/pur/pur5170",
     [("SELECT","getList","MRO 발주현황 조회"),
      ("SELECT","getList","기간/상태 조건 조회")]),

    ("pur_5180M", "MRO 물품발주현황","구매관리", "MRO",        "MRO물품발주현황",
     "/mis/pur/pur1320007",
     [("SELECT","getList","MRO 물품발주 목록 조회")]),

    ("pur_5190M", "MRO물품상세내역", "구매관리", "MRO",        "MRO물품상세내역",
     "/mis/pur/pur1330007",
     [("SELECT","getList","MRO 물품 상세 내역 조회")]),

    # ── 입찰/규격/예가 업무 ──────────────────────────────────────────
    ("pur_8010M", "규격입찰관리",    "구매관리", "규격입찰",   "규격입찰관리",
     "/mis/pur/pur8010",
     [("SELECT","getList","규격입찰 목록 조회"),
      ("SELECT","getList","기간/상태 조건 조회")]),

    ("pur_8011M", "규격입찰결과등록","구매관리", "규격입찰",   "규격입찰결과등록",
     "/mis/pur/pur8011",
     [("SELECT","getData","규격입찰결과 조회"),
      ("INSERT","setData","규격입찰결과 등록"),
      ("DELETE","delData","규격입찰결과 삭제"),
      ("UPDATE","setAppro","규격입찰 승인"),
      ("UPDATE","setReject","규격입찰 반려"),
      ("UPDATE","setAccp","규격입찰 접수")]),

    ("pur_8020M", "개찰결과관리",    "구매관리", "개찰",       "개찰결과관리",
     "/mis/pur/pur8020",
     [("SELECT","getList","개찰결과 목록 조회"),
      ("SELECT","getList","기간 조건 조회")]),

    ("pur_8021M", "개찰결과등록",    "구매관리", "개찰",       "개찰결과등록",
     "/mis/pur/pur8021",
     [("SELECT","getData","개찰결과 상세 조회"),
      ("INSERT","setData","개찰결과 등록"),
      ("DELETE","delData","개찰결과 삭제"),
      ("UPDATE","setAppro","개찰결과 승인"),
      ("UPDATE","setAccp","개찰결과 접수")]),

    ("pur_8023M", "개찰결과조회",    "구매관리", "개찰",       "개찰결과조회",
     "/mis/pur/pur0110",
     [("SELECT","getList","개찰결과 조회"),
      ("SELECT","getList2","직접구매 개찰결과 조회")]),

    ("pur_8030M", "학술연구용품관리","구매관리", "학술연구",   "학술연구용품관리",
     "/mis/pur/pur8030",
     [("SELECT","getList","학술연구용품 목록 조회"),
      ("SELECT","getList","기간 조건 조회")]),

    ("pur_8031M", "학술연구용품신청","구매관리", "학술연구",   "학술연구용품신청",
     "/mis/pur/pur8031",
     [("SELECT","getData","학술연구용품 상세 조회"),
      ("INSERT","setData","학술연구용품 신청 등록"),
      ("UPDATE","setData","학술연구용품 신청 수정"),
      ("DELETE","delData","학술연구용품 신청 삭제")]),

    ("pur_8040M", "예정가격조사관리","구매관리", "예정가격",   "예정가격조사관리",
     "/mis/pur/pur8040",
     [("SELECT","getList","예정가격조사 목록 조회"),
      ("SELECT","getList","기간 조건 조회")]),

    ("pur_8041M", "예정가격조사등록","구매관리", "예정가격",   "예정가격조사등록",
     "/mis/pur/pur8041",
     [("SELECT","getData","예정가격조사 상세 조회"),
      ("INSERT","setData","예정가격조사 신규 등록"),
      ("DELETE","delData","예정가격조사 삭제"),
      ("UPDATE","setAccp","예정가격 접수"),
      ("UPDATE","setAppro","예정가격 승인"),
      ("UPDATE","setReject","예정가격 반려")]),

    ("pur_8050M", "사전규격",        "구매관리", "사전규격",   "사전규격",
     "/mis/pur/pur8050",
     [("SELECT","getList","사전규격 목록 조회"),
      ("SELECT","getList","기간 조건 조회")]),

    ("pur_8051M", "사전규격결과등록","구매관리", "사전규격",   "사전규격결과등록",
     "/mis/pur/pur8051",
     [("SELECT","getData","사전규격결과 조회"),
      ("INSERT","setData","사전규격결과 등록"),
      ("DELETE","delData","사전규격결과 삭제"),
      ("UPDATE","setAccp","사전규격 접수"),
      ("UPDATE","setAppro","사전규격 승인"),
      ("UPDATE","setReject","사전규격 반려")]),

    ("pur_9999M", "업체관리",        "구매관리", "기준정보",   "업체관리",
     "/mis/pur/pur9999",
     [("SELECT","getList","업체 목록 조회"),
      ("SELECT","getList2","업체 추가 정보 조회"),
      ("INSERT","setData","신규 업체 등록"),
      ("UPDATE","setData","업체 정보 수정")]),
]


# ─── 시나리오 세부 생성 함수 ────────────────────────────────────────────────
def make_scenarios(scr):
    screen_id, screen_nm, mid, sub, menu, base_url, cruds = scr
    screen_code = screen_id.replace("pur_", "").replace("M", "").upper()

    rows = []

    for seq, (crud, url_sfx, scen_sfx) in enumerate(cruds, start=1):
        # ── 통합 테스트 시나리오 ──
        it_id = f"IT_PUR{screen_code}_{seq:04d}"
        it_url = f"{base_url}/{url_sfx}.do"

        if crud == "SELECT" and seq == 1:
            flow = (f"메뉴 이동 > [{menu}] 화면 진입 > "
                    f"초기 {scen_sfx.replace('조회','').strip()} 조회 확인")
            inputs = "없음 (자동 조회)"
            expected = f"{scen_sfx} 목록 1건 이상 조회됨"
        elif crud == "SELECT" and seq > 1:
            flow = (f"메뉴 이동 > [{menu}] 화면 진입 > "
                    f"검색조건 입력 > [조회] 버튼 클릭 > 결과 확인")
            inputs = "시작일=TODAY-90; 종료일=TODAY; 상태코드=(전체)"
            expected = f"조건에 맞는 {sub} 목록 0건 이상 조회됨"
        elif crud == "INSERT":
            flow = (f"메뉴 이동 > [{menu}] 화면 진입 > "
                    f"[신규] 또는 [행추가] 버튼 클릭 > 필수항목 입력 > "
                    f"[저장] 버튼 클릭 > 저장 확인 다이얼로그 [확인] > 목록 확인")
            inputs = f"필수항목=TEST_{screen_code}; 작성일=TODAY"
            expected = f"저장 완료 후 목록에 등록된 {sub} 1건 이상 조회됨"
        elif crud == "UPDATE":
            flow = (f"메뉴 이동 > [{menu}] 화면 진입 > "
                    f"목록에서 대상 선택 > 수정 항목 변경 > "
                    f"[저장] 버튼 클릭 > 저장 확인 > 변경 내용 확인")
            inputs = f"수정항목=수정_TEST_{screen_code}"
            expected = "수정 완료 후 변경된 내용이 목록에 반영됨"
        elif crud == "DELETE":
            flow = (f"메뉴 이동 > [{menu}] 화면 진입 > "
                    f"삭제 대상 행 선택 > [삭제] 버튼 클릭 > "
                    f"삭제 확인 다이얼로그 [확인] > 삭제 결과 확인")
            inputs = "삭제대상=등록된 테스트 데이터"
            expected = "삭제 완료 후 해당 항목이 목록에서 제거됨"
        else:
            flow = (f"메뉴 이동 > [{menu}] 화면 진입 > "
                    f"대상 선택 > {url_sfx} 처리 > 결과 확인")
            inputs = "대상항목=처리 대상 데이터"
            expected = f"{url_sfx} 처리 완료 후 상태 변경 반영됨"

        rows.append({
            "구분": "통합",
            "시나리오ID": it_id,
            "화면ID": screen_id,
            "화면명": screen_nm,
            "중분류": mid,
            "소분류": sub,
            "메뉴명": menu,
            "CRUD": crud,
            "API URL": it_url,
            "시나리오명": f"[{crud}] {scen_sfx}",
            "시나리오흐름": flow,
            "입력값": inputs,
            "예상결과": expected,
        })

        # ── 단위 테스트 시나리오 (SELECT & INSERT/UPDATE/DELETE 각 1개) ──
        if crud in ("SELECT",) and seq == 1 or crud == "INSERT" or crud == "UPDATE" or crud == "DELETE":
            ut_id = f"UT_PUR{screen_code}_{seq:04d}"
            ut_expected_map = {
                "SELECT": f"HTTP 200, {sub} 목록 데이터 1건 이상 포함",
                "INSERT": "HTTP 200, resultCnt≥1 또는 result='SUCCESS'",
                "UPDATE": "HTTP 200, resultCnt≥1 또는 result='SUCCESS'",
                "DELETE": "HTTP 200, resultCnt≥1 또는 result='SUCCESS'",
            }
            ut_input_map = {
                "SELECT":  "pageIndex=1; pageUnit=10; 검색조건=(기본값)",
                "INSERT":  f"등록 대상 DTO (필수항목 포함): 항목코드=TEST_{screen_code}",
                "UPDATE":  f"수정 대상 DTO: 수정항목=수정값",
                "DELETE":  "삭제 대상 키값 (등록된 테스트 데이터)",
            }
            rows.append({
                "구분": "단위",
                "시나리오ID": ut_id,
                "화면ID": screen_id,
                "화면명": screen_nm,
                "중분류": mid,
                "소분류": sub,
                "메뉴명": menu,
                "CRUD": crud,
                "API URL": it_url,
                "시나리오명": f"[단위][{crud}] {scen_sfx} API",
                "시나리오흐름": (f"POST {it_url} 직접 호출 > "
                               f"응답 코드 200 확인 > 응답 데이터 검증"),
                "입력값": ut_input_map.get(crud, ""),
                "예상결과": ut_expected_map.get(crud, "HTTP 200"),
            })

    return rows


# ─── Excel 생성 ─────────────────────────────────────────────────────────────
def build_excel(out_path):
    wb = openpyxl.Workbook()
    ws_all   = wb.active
    ws_all.title = "전체시나리오"
    ws_integ = wb.create_sheet("통합테스트")
    ws_unit  = wb.create_sheet("단위테스트")

    def apply_header(ws):
        ws.append(HEADERS)
        for ci, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=ci)
            cell.font      = Font(name="맑은 고딕", bold=True,
                                  color=HEADER_FONT, size=10)
            cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center", wrap_text=True)
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[ci-1]
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    for ws in (ws_all, ws_integ, ws_unit):
        apply_header(ws)

    all_rows, integ_rows, unit_rows = [], [], []
    seq = 0
    for scr in PUR_SCREENS:
        scen_rows = make_scenarios(scr)
        all_rows.extend(scen_rows)
        integ_rows.extend(r for r in scen_rows if r["구분"] == "통합")
        unit_rows.extend(r for r in scen_rows if r["구분"] == "단위")

    def write_rows(ws, rows):
        prev_screen = None
        row_idx = 2
        no = 0
        for r in rows:
            # 화면 구분자 행
            if r["화면ID"] != prev_screen:
                prev_screen = r["화면ID"]
                sep_cell = ws.cell(row=row_idx, column=1,
                                   value=f"▶ {r['화면ID']}  {r['화면명']}")
                sep_cell.font  = Font(name="맑은 고딕", bold=True,
                                      color=SEP_FONT, size=9)
                sep_cell.fill  = PatternFill("solid", fgColor=SEP_BG)
                sep_cell.alignment = Alignment(vertical="center")
                ws.merge_cells(start_row=row_idx, start_column=1,
                               end_row=row_idx, end_column=len(HEADERS))
                ws.row_dimensions[row_idx].height = 16
                row_idx += 1

            no += 1
            is_integ = (r["구분"] == "통합")
            bg = INTEG_BG if is_integ else UNIT_BG
            fill = PatternFill("solid", fgColor=bg)

            vals = [
                no,
                r["구분"],
                r["시나리오ID"],
                r["화면ID"],
                r["화면명"],
                r["중분류"],
                r["소분류"],
                r["메뉴명"],
                r["CRUD"],
                r["API URL"],
                r["시나리오명"],
                r["시나리오흐름"],
                r["입력값"],
                r["예상결과"],
                "",  # 테스트결과
                "",  # 실패사유
                "",  # 비고
            ]
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.font      = Font(name="맑은 고딕", size=9)
                cell.fill      = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True,
                                           horizontal="left")
                cell.border    = BORDER
            ws.row_dimensions[row_idx].height = 40
            row_idx += 1

    write_rows(ws_all,   all_rows)
    write_rows(ws_integ, integ_rows)
    write_rows(ws_unit,  unit_rows)

    # ── 요약 시트 ─────────────────────────────────────────────────
    ws_sum = wb.create_sheet("요약", 0)
    ws_all = wb.worksheets[1]  # 재참조

    sum_headers = ["화면ID", "화면명", "중분류", "소분류",
                   "통합_전체", "통합_SELECT", "통합_INSERT",
                   "통합_UPDATE", "통합_DELETE", "통합_기타",
                   "단위_전체", "단위_SELECT", "단위_INSERT",
                   "단위_UPDATE", "단위_DELETE"]
    SUM_WIDTHS  = [14, 20, 12, 14, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]
    ws_sum.append(sum_headers)
    for ci, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(1, ci)
        c.font = Font(name="맑은 고딕", bold=True, color=HEADER_FONT, size=9)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws_sum.column_dimensions[get_column_letter(ci)].width = SUM_WIDTHS[ci-1]
    ws_sum.row_dimensions[1].height = 22
    ws_sum.freeze_panes = "A2"

    from collections import defaultdict
    stat = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for r in all_rows:
        sid  = r["화면ID"]
        gbn  = r["구분"]
        crud = r["CRUD"]
        stat[sid][gbn][crud] += 1
        stat[sid][gbn]["전체"] += 1

    sr = 2
    for scr in PUR_SCREENS:
        sid, snm, mid, sub = scr[0], scr[1], scr[2], scr[3]
        s = stat[sid]
        row_vals = [
            sid, snm, mid, sub,
            s["통합"]["전체"],
            s["통합"]["SELECT"], s["통합"]["INSERT"],
            s["통합"]["UPDATE"], s["통합"]["DELETE"],
            s["통합"]["전체"] - s["통합"]["SELECT"] - s["통합"]["INSERT"]
                               - s["통합"]["UPDATE"] - s["통합"]["DELETE"],
            s["단위"]["전체"],
            s["단위"]["SELECT"], s["단위"]["INSERT"],
            s["단위"]["UPDATE"], s["단위"]["DELETE"],
        ]
        for ci, v in enumerate(row_vals, 1):
            c = ws_sum.cell(sr, ci, value=v)
            c.font = Font(name="맑은 고딕", size=9)
            c.alignment = Alignment(horizontal="center" if ci > 4 else "left",
                                    vertical="center")
            c.border = BORDER
        sr += 1

    # 합계 행
    tot_row = ["합계", f"{len(PUR_SCREENS)}개 화면", "", ""]
    it_cnt  = sum(1 for r in all_rows if r["구분"]=="통합")
    ut_cnt  = sum(1 for r in all_rows if r["구분"]=="단위")
    for v in [it_cnt,
              sum(1 for r in all_rows if r["구분"]=="통합" and r["CRUD"]=="SELECT"),
              sum(1 for r in all_rows if r["구분"]=="통합" and r["CRUD"]=="INSERT"),
              sum(1 for r in all_rows if r["구분"]=="통합" and r["CRUD"]=="UPDATE"),
              sum(1 for r in all_rows if r["구분"]=="통합" and r["CRUD"]=="DELETE"),
              it_cnt - sum(1 for r in all_rows if r["구분"]=="통합" and r["CRUD"] in ("SELECT","INSERT","UPDATE","DELETE")),
              ut_cnt,
              sum(1 for r in all_rows if r["구분"]=="단위" and r["CRUD"]=="SELECT"),
              sum(1 for r in all_rows if r["구분"]=="단위" and r["CRUD"]=="INSERT"),
              sum(1 for r in all_rows if r["구분"]=="단위" and r["CRUD"]=="UPDATE"),
              sum(1 for r in all_rows if r["구분"]=="단위" and r["CRUD"]=="DELETE")]:
        tot_row.append(v)
    for ci, v in enumerate(tot_row, 1):
        c = ws_sum.cell(sr, ci, value=v)
        c.font = Font(name="맑은 고딕", bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A3C72")
        c.alignment = Alignment(horizontal="center" if ci > 2 else "left",
                                vertical="center")
        c.border = BORDER

    wb.save(out_path)
    return all_rows, integ_rows, unit_rows


# ─── MAIN ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    out = r"<workspace-root>/시나리오_1차.xlsx"
    all_r, it_r, ut_r = build_excel(out)
    print(f"✅ 생성 완료: {out}")
    print(f"   전체 시나리오: {len(all_r)}건")
    print(f"   통합 테스트:   {len(it_r)}건")
    print(f"   단위 테스트:   {len(ut_r)}건")
    print(f"   대상 화면:     {len(PUR_SCREENS)}개")
