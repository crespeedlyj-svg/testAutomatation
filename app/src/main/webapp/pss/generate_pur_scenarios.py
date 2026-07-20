#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PUR 모듈 테스트 시나리오 생성기 v3
==============================================
개선 사항 (v3):
  ★ scenario-rules-pur.md 를 파싱하여 규칙을 동적으로 로드
    - API 호출 없음, 순수 텍스트 파싱
    - md 파일이 없으면 내부 기본값으로 자동 fallback
    - md 수정 → Python 재실행만 하면 즉시 반영

  이전 개선사항 유지:
  1. ServiceImpl.java 예외 패턴 파싱
  2. APV_STAT_CD / PUR_STEP 상태 머신
  3. 화면 간 데이터 흐름 의존성
  4. SQLMAP INSERT 컬럼 기반 테스트 데이터

사용법:
    python generate_pur_scenarios.py [프로젝트루트경로]
출력:
    pur_test_scenarios_YYYYMMDD.xlsx
"""

import os, re, sys
from pathlib import Path
from datetime import date
from collections import defaultdict


# ══════════════════════════════════════════════════════════════════════════════
# 0. scenario-rules-pur.md 파서  (AI API 호출 없음 — 순수 텍스트 파싱)
# ══════════════════════════════════════════════════════════════════════════════

def _parse_md_table(section_text: str) -> list:
    """
    마크다운 표를 파싱하여 행 목록 반환.
    각 행은 컬럼 값의 리스트.  구분선(---|---) 행은 제외.
    """
    rows = []
    for line in section_text.splitlines():
        line = line.strip()
        if not line.startswith('|'):
            continue
        cells = [c.strip() for c in line.split('|') if c.strip()]
        if not cells:
            continue
        # 구분선 행 (---|---) 제외
        if all(re.match(r'^[-:]+$', c) for c in cells):
            continue
        rows.append(cells)
    return rows   # rows[0] = 헤더, rows[1:] = 데이터


def _get_section(md_text: str, header: str) -> str:
    """
    ## {header} 섹션의 텍스트를 반환.
    다음 ## 섹션 전까지의 내용.
    """
    pattern = re.compile(
        r'##\s+' + re.escape(header) + r'\s*\n(.*?)(?=\n##\s|\Z)',
        re.DOTALL | re.IGNORECASE
    )
    m = pattern.search(md_text)
    return m.group(1) if m else ''


def load_rules_from_md(md_path: Path) -> dict:
    """
    scenario-rules-pur.md 를 읽어서 규칙 딕셔너리 반환.
    파일이 없거나 파싱 실패 시 빈 딕셔너리 반환 → 기본값 사용.
    """
    rules = {}
    if not md_path.exists():
        print(f'[INFO] rules 파일 없음 ({md_path.name}) → 내부 기본값 사용')
        return rules

    try:
        text = md_path.read_text(encoding='utf-8', errors='replace')
        print(f'[INFO] rules 로드: {md_path.name}')
    except Exception as e:
        print(f'[WARN] rules 읽기 실패: {e} → 내부 기본값 사용')
        return rules

    # ── 섹션 2.1: 결재 상태 (APV_STAT_CD) ────────────────────────────────────
    # 표 형식: | 코드 | 명칭 | 허용 작업 | 차단 작업 |
    sec_apv = _get_section(text, '2. 핵심 상태 코드')
    # 2.1 소섹션만 추출
    apv_m = re.search(r'###\s*2\.1[^\n]*\n(.*?)(?=###|\Z)', sec_apv, re.DOTALL)
    if apv_m:
        rows = _parse_md_table(apv_m.group(1))
        state_machine = {}
        for row in rows[1:]:           # 헤더 제외
            if len(row) < 4:
                continue
            code, name, allowed_raw, blocked_raw = row[0], row[1], row[2], row[3]
            # 중간 점(·) 구분 목록 파싱
            allowed = [a.strip() for a in re.split(r'[·,，]', allowed_raw) if a.strip() and a.strip() != '없음']
            blocked = [b.strip() for b in re.split(r'[·,，]', blocked_raw) if b.strip() and b.strip() != '없음']
            label = f'{name}({code.replace("000-010-", "010-")})'
            state_machine[label] = {
                'allowed':   allowed,
                'blocked':   blocked,
                'test_data': f"APV_STAT_CD='{code}'",
            }
        if state_machine:
            rules['state_machine'] = state_machine
            print(f'       state_machine: {len(state_machine)}개 상태 로드')

    # ── 섹션 2.2: PUR 단계 코드 ──────────────────────────────────────────────
    # 표 형식: | 코드 | 명칭 | 진입 조건 |
    pur_step_m = re.search(r'###\s*2\.2[^\n]*\n(.*?)(?=###|##|\Z)', sec_apv, re.DOTALL)
    if not pur_step_m:
        pur_step_m = re.search(r'###\s*2\.2[^\n]*\n(.*?)(?=###|##|\Z)', text, re.DOTALL)
    if pur_step_m:
        rows = _parse_md_table(pur_step_m.group(1))
        pur_step = {}
        for row in rows[1:]:
            if len(row) >= 2:
                pur_step[row[0]] = row[1]
        if pur_step:
            rules['pur_step'] = pur_step
            print(f'       pur_step: {len(pur_step)}개 단계 로드')

    # ── 섹션 5: 필드 테스트 데이터 ───────────────────────────────────────────
    # 표 형식: | 필드 패턴 | 테스트 값 | 비고 |
    sec5 = _get_section(text, '5. 필드 테스트 데이터 표준')
    if not sec5:
        sec5 = _get_section(text, '5.')
    rows5 = _parse_md_table(sec5)
    field_values = {}
    for row in rows5[1:]:
        if len(row) >= 2:
            field_pattern = row[0].strip('`')
            test_val      = row[1].strip('`').strip()
            field_values[field_pattern] = test_val
    if field_values:
        rules['field_values'] = field_values
        print(f'       field_values: {len(field_values)}개 필드 로드')

    # ── 섹션 6: CRUD별 예상결과 문구 ─────────────────────────────────────────
    # 표 형식: | CRUD | 예상결과 문구 |
    sec6 = _get_section(text, '6. CRUD별 예상결과 표준 문구')
    if not sec6:
        sec6 = _get_section(text, '6.')
    rows6 = _parse_md_table(sec6)
    crud_expected = {}
    for row in rows6[1:]:
        if len(row) >= 2:
            crud_key  = row[0].strip().split('/')[0].strip()   # "SAVE/INSERT" → "SAVE"
            crud_text = row[1].strip()
            # {화면명} 플레이스홀더는 Python에서 .format(화면명=...) 으로 치환
            crud_expected[crud_key] = crud_text
            # "SAVE/INSERT" 처리
            for alias in row[0].split('/'):
                crud_expected[alias.strip()] = crud_text
    if crud_expected:
        rules['crud_expected'] = crud_expected
        print(f'       crud_expected: {len(crud_expected)}개 CRUD 문구 로드')

    # ── 섹션 4.3: 비정상 시나리오 유형 목록 ──────────────────────────────────
    # "#### ①" 형식의 소항목 파싱
    sec43 = re.search(r'###\s*4\.3[^\n]*\n(.*?)(?=###\s*4\.4|###\s*5\.|##\s*5\.|\Z)',
                      text, re.DOTALL)
    neg_types = []
    if sec43:
        for m in re.finditer(r'####[^#\n]*?(①|②|③|④|⑤)[^\n]*\n([^\n]+)', sec43.group(1)):
            neg_types.append(m.group(2).strip())
    if neg_types:
        rules['neg_types'] = neg_types
        print(f'       neg_types: {len(neg_types)}개 비정상 유형 로드')

    # ── 섹션 7: 원인행위 관련 규칙 ───────────────────────────────────────────
    sec7 = _get_section(text, '7. 원인행위')
    if not sec7:
        sec7 = _get_section(text, '7.')
    if 'setPurControlList' in sec7 or '원인행위' in sec7:
        rules['has_control_act_rule'] = True

    total = sum(len(v) if isinstance(v, dict) else 1 for v in rules.values())
    print(f'[INFO] rules 총 {total}개 항목 로드 완료')
    return rules

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ══════════════════════════════════════════════════════════════════════════════
# 1. PUR 도메인 상수 (기본값 — md 로드 후 오버라이드됨)
# ══════════════════════════════════════════════════════════════════════════════

# ── 결재 상태 코드 (APV_STAT_CD) — 기본값 ────────────────────────────────────
APV_STAT = {
    "000-010-010": "임시저장",
    "000-010-020": "결재요청",
    "000-010-030": "결재진행",
    "000-010-040": "결재완료",
    "000-010-050": "반려",
    "000-010-060": "회수",
    "000-010-090": "전체(진행중 포함)",
}

# ── 구매 단계 코드 (PUR_STEP) — 기본값 ──────────────────────────────────────
PUR_STEP = {
    "601-010": "구매요청",   "601-011": "구매취소",
    "601-020": "구매심사",   "601-030": "계약준비",
    "601-040": "계약중",     "601-050": "발주",
    "601-060": "납품",       "601-070": "지출(정산전)",
    "601-080": "정산완료",   "601-090": "계약해지",
    "601-091": "계약취소",
}

# ── 상태 머신 — 기본값 ────────────────────────────────────────────────────────
STATE_MACHINE = {
    "임시저장(010-010)": {
        "allowed":  ["getData(조회)", "setData(저장)", "delData(삭제)", "결재요청(setApp)"],
        "blocked":  [],
        "test_data": "APV_STAT_CD='000-010-010'",
    },
    "결재요청(010-020)": {
        "allowed":  ["getData(조회)", "결재취소(회수)"],
        "blocked":  ["setData(수정)", "delData(삭제)"],
        "test_data": "APV_STAT_CD='000-010-020'",
    },
    "결재진행(010-030)": {
        "allowed":  ["getData(조회)"],
        "blocked":  ["setData(수정)", "delData(삭제)", "결재요청"],
        "test_data": "APV_STAT_CD='000-010-030'",
    },
    "결재완료(010-040)": {
        "allowed":  ["getData(조회)", "다음단계 진행"],
        "blocked":  ["setData(수정)", "delData(삭제)"],
        "test_data": "APV_STAT_CD='000-010-040'",
    },
    "반려(010-050)": {
        "allowed":  ["getData(조회)", "setData(수정)", "결재재요청"],
        "blocked":  [],
        "test_data": "APV_STAT_CD='000-010-050'",
    },
}

# ── CRUD 예상결과 문구 — 기본값 ───────────────────────────────────────────────
_CRUD_EXPECTED_DEFAULT = {
    "SELECT":   "{screen_desc} 조회 조건에 맞는 목록/상세가 정상 반환된다.",
    "SAVE":     "{screen_desc} 데이터가 저장되고 키값(RQST_NO 등)이 반환된다. 저장 후 목록에 즉시 반영된다.",
    "INSERT":   "{screen_desc} 데이터가 등록되고 목록에 반영된다.",
    "UPDATE":   "수정 내용이 반영되고 수정 전 데이터가 더 이상 조회되지 않는다.",
    "DELETE":   "선택 건이 삭제되고 목록에서 제거된다. 삭제 후 재조회 시 해당 건이 없어야 한다.",
    "APPROVAL": "결재 요청 완료 후 APV_STAT_CD가 결재요청(000-010-020)으로 변경된다.",
    "CANCEL":   "취소 처리 후 PUR_STEP이 취소 코드로 변경되고 관련 예산/원인행위가 복원된다.",
    "CHECK":    "검증 결과(CHK_OK=Y/N)와 상세 메시지(CHK_MSG)가 반환된다.",
}

# ── 필드 테스트 값 — 기본값 ──────────────────────────────────────────────────
_FIELD_VALUES_DEFAULT = {
    "RQST_SBJ":        "테스트구매요청_화면번호",
    "RQST_DT":         date.today().strftime('%Y%m%d'),
    "DLVERY_SDT":      date.today().strftime('%Y%m%d'),
    "DLVERY_EDT":      date.today().strftime('%Y%m%d'),
    "TOT_PRES_AMT":    "1000000",
    "TOT_PRES_AMT_부족": "99999999",
    "PUR_TP":          "602-001",
    "FRGN_CLS":        "604-001",
    "VAT_TP":          "1",
    "APV_STAT_CD_임시": "000-010-010",
    "APV_STAT_CD_완료": "000-010-040",
    "CORP_CD":         "01",
}

# ── md 규칙이 적용될 전역 컨테이너 ───────────────────────────────────────────
# generate_scenarios() 호출 전에 apply_rules()로 채워짐
_RULES: dict = {}


def apply_rules(rules: dict):
    """
    load_rules_from_md()의 결과를 전역 상수에 적용한다.
    md에 해당 섹션이 없으면 기존 기본값을 그대로 유지한다.
    """
    global PUR_STEP, STATE_MACHINE, _RULES
    _RULES = rules

    if rules.get('pur_step'):
        PUR_STEP.update(rules['pur_step'])
        # ServiceImpl 파서에서도 사용하므로 전역 갱신
        print(f'[APPLY] PUR_STEP: {len(PUR_STEP)}개')

    if rules.get('state_machine'):
        STATE_MACHINE.update(rules['state_machine'])
        print(f'[APPLY] STATE_MACHINE: {len(STATE_MACHINE)}개 상태')

# ── 화면 간 데이터 흐름 의존성 ────────────────────────────────────────────────
# 이 화면을 사용하려면 선행 화면이 완료(결재완료)되어야 함
SCREEN_FLOW = {
    "0010": {
        "name": "구매요청",
        "table": "PUR_RQST_MST",
        "key": "RQST_NO",
        "required_fields": ["RQST_SBJ", "RQST_DT", "PUR_TP", "DLVERY_SDT",
                            "APNT_EMP_NO", "APNT_DEPT_CD", "FRGN_CLS"],
        "required_amount_fields": ["TOT_PRES_AMT"],
        "pre_screen": None,
        "pre_key": None,
        "pre_stat": None,
        "pur_step_after": "601-010",
    },
    "0140": {
        "name": "구매요구현황",
        "table": "PUR_RQST_MST",
        "key": "RQST_NO",
        "required_fields": ["RQST_NO"],
        "required_amount_fields": [],
        "pre_screen": "0010",
        "pre_key": "RQST_NO",
        "pre_stat": "000-010-040",
        "pur_step_after": "601-020",
    },
    "0150": {
        "name": "구매계약목록",
        "table": "PUR_STEP_MGT_MST",
        "key": "RGST_NO",
        "required_fields": ["RGST_NO"],
        "required_amount_fields": [],
        "pre_screen": "0140",
        "pre_key": "RQST_NO",
        "pre_stat": "000-010-040",
        "pur_step_after": "601-030",
    },
    "0153": {
        "name": "구매계약상세",
        "table": "COM_RQST_NO",
        "key": "RQST_NO",
        "required_fields": ["RQST_NO", "RQST_SBJ", "DLVERY_SDT", "DLVERY_EDT",
                            "TOT_PRES_AMT", "VAT_TP", "PUR_TP", "FRGN_CLS"],
        "required_amount_fields": ["TOT_PRES_AMT"],
        "pre_screen": "0150",
        "pre_key": "RGST_NO",
        "pre_stat": "000-010-040",
        "pur_step_after": "601-040",
        "budget_check": True,
        "concurrency_check": True,
    },
    "0200": {
        "name": "계약대상목록(발주)",
        "table": "PUR_STEP_MGT_MST",
        "key": "RGST_NO",
        "required_fields": ["RGST_NO"],
        "required_amount_fields": [],
        "pre_screen": "0153",
        "pre_key": "RQST_NO",
        "pre_stat": "000-010-040",
        "pur_step_after": "601-050",
    },
    "0300": {
        "name": "납품등록",
        "table": "PUR_DLVR_MST",
        "key": "DLVR_NO",
        "required_fields": ["DLVR_DT", "DLVR_AMT", "DLVR_EMP_NO"],
        "required_amount_fields": ["DLVR_AMT"],
        "pre_screen": "0200",
        "pre_key": "RGST_NO",
        "pre_stat": "000-010-040",
        "pur_step_after": "601-060",
    },
    "0400": {
        "name": "대가지급등록",
        "table": "PUR_PAY_MST",
        "key": "PAY_NO",
        "required_fields": ["PAY_DT", "PAY_AMT", "ACNT_NO", "BANK_CD"],
        "required_amount_fields": ["PAY_AMT"],
        "pre_screen": "0300",
        "pre_key": "DLVR_NO",
        "pre_stat": "000-010-040",
        "pur_step_after": "601-070",
    },
    "0500": {
        "name": "구매정산",
        "table": "PUR_CTRCT_MST",
        "key": "CTRCT_NO",
        "required_fields": ["CTRCT_NO"],
        "required_amount_fields": [],
        "pre_screen": "0400",
        "pre_key": "PAY_NO",
        "pre_stat": "000-010-040",
        "pur_step_after": "601-080",
    },
}

# ── URL suffix → CRUD 분류 ────────────────────────────────────────────────────
URL_CRUD = {
    "getlist": "SELECT", "getdata": "SELECT", "getdetaildata": "SELECT",
    "getctrctlist": "SELECT", "getbudgdata": "SELECT", "getcombo": "SELECT",
    "getrqstmst": "SELECT", "getmst": "SELECT", "getdtl": "SELECT",
    "getitem": "SELECT", "getbudgbaln": "SELECT", "getmstlist": "SELECT",
    "getrqstlist": "SELECT", "getordlist": "SELECT", "getdlvrlist": "SELECT",
    "getpaylist": "SELECT", "getsrchlist": "SELECT",
    "setdata": "SAVE", "setsave": "SAVE", "setapp": "APPROVAL",
    "setappstat": "APPROVAL", "setcancel": "CANCEL", "setconfirm": "SAVE",
    "setcomplete": "SAVE", "setrequest": "SAVE", "setregist": "SAVE",
    "deldata": "DELETE", "dellist": "DELETE", "delitem": "DELETE",
    "uptchk": "CHECK", "checkbudgbaln": "CHECK", "rqsttpchk": "CHECK",
    "getreslenblYn": "CHECK",
}

# ── 화면 번호 → 카테고리 ─────────────────────────────────────────────────────
SCREEN_CATEGORY = {
    "0000": ("목록", "구매 전체 목록"),
    "0010": ("구매요청", "구매요청서 등록/수정"),
    "0011": ("구매요청", "구매요청 상세"),
    "0043": ("구매요청", "구매요청 일괄처리"),
    "0110": ("구매계획", "구매계획 수립"),
    "0111": ("구매계획", "구매계획 상세"),
    "0112": ("구매계획", "구매계획 변경"),
    "0121": ("구매계획", "구매계획 확정"),
    "0131": ("구매계획", "구매계획 취소"),
    "0140": ("구매요구현황", "구매요구 현황 조회"),
    "0141": ("구매요구현황", "구매요구 상세"),
    "0150": ("구매계약목록", "구매계약 목록"),
    "0151": ("구매계약", "구매계약 상세"),
    "0153": ("구매계약", "구매계약 상세(예산체크·동시성)"),
    "0200": ("계약대상/발주", "발주 대상 계약 목록"),
    "0210": ("발주", "발주 상세"),
    "0211": ("발주", "발주 변경"),
    "0221": ("발주", "발주 확정"),
    "0300": ("납품/검수", "납품 등록"),
    "0311": ("납품/검수", "납품 상세"),
    "0321": ("납품/검수", "검수 등록"),
    "0330": ("납품/검수", "검수 완료"),
    "0350": ("납품/검수", "납품 현황"),
    "0351": ("납품/검수", "납품 취소"),
    "0400": ("대가지급", "대가지급 등록"),
    "0411": ("대가지급", "지급 상세"),
    "0420": ("대가지급", "지급 처리"),
    "0421": ("대가지급", "지급 취소"),
    "0430": ("대가지급", "세금계산서 처리"),
    "0431": ("대가지급", "세금계산서 상세"),
    "0440": ("대가지급", "대가지급 현황"),
    "0500": ("정산", "구매 정산"),
    "0511": ("정산", "정산 상세"),
    "0522": ("정산", "정산 확정"),
    "0530": ("정산", "정산 현황"),
    "0591": ("정산", "정산 관련"),
    "0610": ("예산", "예산 현황"),
    "0611": ("예산", "예산 상세"),
    "0710": ("업체", "업체 관리"),
    "0711": ("업체", "업체 상세"),
    "5110": ("해외구매", "해외구매 요청"),
    "5120": ("해외구매", "해외계약"),
    "5130": ("해외구매", "해외발주"),
    "5140": ("해외구매", "해외납품"),
    "5150": ("해외구매", "해외대가지급"),
    "5160": ("해외구매", "해외정산"),
    "8010": ("코드관리", "구매 코드"),
    "9002": ("현황조회", "구매 현황"),
    "9004": ("현황조회", "계약 현황"),
    "9401": ("현황조회", "구매 진행 현황"),
}

# ══════════════════════════════════════════════════════════════════════════════
# 2. ServiceImpl 예외 패턴 파서
# ══════════════════════════════════════════════════════════════════════════════

# 추출할 예외 패턴 유형
EXCEPTION_PATTERNS = [
    # (패턴 정규식, 유형, 설명 추출 함수)
    (r'throw\s+new\s+SystemException\s*\(\s*"([^"]+)"',     "SystemException", lambda m: m.group(1)),
    (r'throw\s+new\s+SystemException\s*\(\s*messagesvc\.getMessage\s*\("([^"]+)"\)',
                                                              "SystemException(메시지코드)", lambda m: m.group(1)),
    (r'if\s*\("([^"]+)"\.equals\s*\(\s*[^)]+\)\s*\)\s*\{[^}]*throw',
                                                              "조건부예외(equals)", lambda m: m.group(1)),
    (r'result\s*=\s*"N".*?return\s+result',                   "결과N반환", lambda m: "result=N 반환"),
    (r'rslt\s*=\s*"N".*?return\s+rslt',                       "결과N반환", lambda m: "rslt=N 반환"),
]

# if 조건 → throw/return 패턴
IF_THROW_PATTERN = re.compile(
    r'if\s*\(([^)]{5,120})\)\s*\{?\s*(?:throw\s+new\s+\w+Exception\s*\("([^"]{3,100})"'
    r'|return\s+"N"'
    r'|return\s+rslt\b)',
    re.DOTALL
)

def parse_service_impl(filepath: Path) -> dict:
    """
    ServiceImpl.java 파싱:
    - 모든 public 메서드 시그니처
    - throw new SystemException 패턴
    - if → throw/return "N" 조건
    - PUR_STEP 세팅 값
    """
    try:
        src = filepath.read_text(encoding='utf-8', errors='replace')
    except Exception:
        return {}

    result = {
        'filename': filepath.name,
        'class_name': '',
        'methods': [],       # [{'name': str, 'exceptions': [...], 'conditions': [...], 'pur_steps': [...]}]
        'global_exceptions': [],
        'pur_steps': [],
    }

    # 클래스명
    cm = re.search(r'class\s+(\w+)', src)
    if cm:
        result['class_name'] = cm.group(1)

    # 전역 SystemException (메서드 바깥)
    for m in re.finditer(r'throw\s+new\s+SystemException\s*\(\s*"([^"]+)"', src):
        result['global_exceptions'].append(m.group(1))

    # PUR_STEP 세팅
    for m in re.finditer(r'put\s*\(\s*"PUR_STEP"\s*,\s*"([^"]+)"\s*\)', src):
        code = m.group(1)
        result['pur_steps'].append({
            'code': code,
            'name': PUR_STEP.get(code, code),
        })

    # 메서드별 분석
    method_re = re.compile(
        r'public\s+([\w<>\[\],?\s]+?)\s+(\w+)\s*\(([^)]*)\)\s*throws\s*\w[^{]*\{',
        re.DOTALL
    )
    for mm in method_re.finditer(src):
        method_name = mm.group(2)
        if method_name in ('toString', 'hashCode', 'equals'):
            continue

        # 메서드 바디 추출 (중괄호 깊이 추적)
        start = mm.end()
        depth = 1
        pos = start
        while pos < len(src) and depth > 0:
            if src[pos] == '{':
                depth += 1
            elif src[pos] == '}':
                depth -= 1
            pos += 1
        body = src[start:pos]

        # SystemException 메시지 추출
        exceptions = []
        for em in re.finditer(r'throw\s+new\s+SystemException\s*\(\s*"([^"]+)"', body):
            exceptions.append({'type': 'SystemException', 'msg': em.group(1)})
        for em in re.finditer(r'throw\s+new\s+SystemException\s*\([^"]*messagesvc[^"]*"([^"]+)"', body):
            exceptions.append({'type': 'SystemException(코드)', 'msg': em.group(1)})

        # if → throw 조건 추출
        conditions = []
        for cm2 in IF_THROW_PATTERN.finditer(body):
            cond = cm2.group(1).strip().replace('\n', ' ').replace('\t', ' ')
            cond = re.sub(r'\s+', ' ', cond)
            msg  = cm2.group(2) if cm2.group(2) else '조건 불충족 시 처리 중단'
            conditions.append({'condition': cond[:100], 'consequence': msg})

        # 메서드 내 PUR_STEP 세팅
        method_steps = []
        for sm in re.finditer(r'put\s*\(\s*"PUR_STEP"\s*,\s*"([^"]+)"', body):
            code = sm.group(1)
            method_steps.append({'code': code, 'name': PUR_STEP.get(code, code)})

        # result="N" 반환 패턴
        has_n_return = bool(re.search(r'return\s+"N"', body))

        result['methods'].append({
            'name':        method_name,
            'exceptions':  exceptions,
            'conditions':  conditions,
            'pur_steps':   method_steps,
            'has_n_return': has_n_return,
        })

    return result

# ══════════════════════════════════════════════════════════════════════════════
# 3. Controller + SQLMAP 파서
# ══════════════════════════════════════════════════════════════════════════════

def parse_controller(filepath: Path) -> dict:
    """Controller.java → URL·메서드 목록"""
    try:
        src = filepath.read_text(encoding='utf-8', errors='replace')
    except Exception:
        return {}

    result = {
        'filename': filepath.name,
        'screen_no': extract_screen_no(filepath.name),
        'pgm_id': (lambda b: b if b and b[-1].isalpha() else b + 'M')(
            filepath.name.replace('Controller.java', '').upper()),  # PUR_0010 → PUR_0010M
        'endpoints': [],
    }

    blocks = re.split(r'(?=@RequestMapping)', src)
    for block in blocks:
        url_m = re.search(r'@RequestMapping\s*\(\s*(?:value\s*=\s*)?["\'](.*?)["\']', block)
        if not url_m:
            continue
        url = url_m.group(1).strip()

        method_m = re.search(
            r'(?:public|protected)\s+\w[\w<>\[\],\s]*?\s+(\w+)\s*\(',
            block[url_m.end():url_m.end() + 300])
        method_name = method_m.group(1) if method_m else url.rsplit('/', 1)[-1].replace('.do', '')

        comment = ''
        jd = re.search(r'/\*\*(.*?)\*/', block[:url_m.start()], re.DOTALL)
        if jd:
            comment = re.sub(r'[\s\*]+', ' ', jd.group(1)).strip()
            comment = re.sub(r'@\w+.*', '', comment).strip()[:60]

        result['endpoints'].append({'url': url, 'method': method_name, 'comment': comment})

    return result


def parse_sqlmap(filepath: Path) -> dict:
    """SQLMAP XML → {queryId: CRUD, insertColumns: {queryId: [col,...]}}"""
    try:
        src = filepath.read_text(encoding='utf-8', errors='replace')
    except Exception:
        return {}

    index   = {}
    columns = {}  # queryId → INSERT 컬럼 목록

    for tag in ('select', 'insert', 'update', 'delete'):
        for m in re.finditer(r'<' + tag + r'[^>]*\bid\s*=\s*["\'](\w+)["\']', src, re.IGNORECASE):
            qid = m.group(1).lower()
            index[qid] = tag.upper()
            if tag == 'insert':
                # INSERT INTO TABLE ( ... ) VALUES 컬럼 추출
                pos  = m.end()
                body = src[pos:pos + 2000]
                col_m = re.search(r'\(\s*([\w\s,\n\r\t]+?)\s*\)\s*(?:VALUES|value)', body, re.IGNORECASE)
                if col_m:
                    raw_cols = col_m.group(1)
                    cols = [c.strip() for c in raw_cols.split(',') if re.match(r'^[A-Z_]+$', c.strip())]
                    if cols:
                        columns[qid] = cols

    return {'crud': index, 'insert_cols': columns}


def extract_screen_no(filename: str) -> str:
    m = re.search(r'[Pp]ur[_]?(\d+)', filename)
    return m.group(1) if m else ''


def classify_crud(url: str, method_name: str, sqlmap_index: dict) -> str:
    mn = method_name.lower()
    if mn in sqlmap_index:
        return sqlmap_index[mn]
    seg = url.rsplit('/', 1)[-1].replace('.do', '').lower()
    if seg in URL_CRUD:
        return URL_CRUD[seg]
    if mn.startswith(('get', 'list', 'search', 'find', 'fetch', 'select', 'read')):
        return 'SELECT'
    if mn.startswith(('del', 'remove', 'purge')):
        return 'DELETE'
    if mn.startswith(('set', 'save', 'insert', 'create', 'add', 'reg')):
        return 'SAVE'
    if mn.startswith(('upt', 'update', 'modify')):
        return 'UPDATE'
    if mn.startswith(('chk', 'check', 'valid', 'rqst')):
        return 'CHECK'
    if 'app' in mn:
        return 'APPROVAL'
    if 'cancel' in mn:
        return 'CANCEL'
    return 'SELECT'

# ══════════════════════════════════════════════════════════════════════════════
# 4. 테스트 데이터 빌더 (SQLMAP INSERT 컬럼 기반)
# ══════════════════════════════════════════════════════════════════════════════

TODAY = date.today().strftime('%Y%m%d')
YEAR_START = date.today().strftime('%Y') + '0101'

# 필드명 패턴 → 기본 테스트 값
FIELD_VALUE_MAP = [
    (r'.*_NO$',      lambda n: 'TEST_' + n[:8]),
    (r'.*_SDT$',     lambda _: YEAR_START),
    (r'.*_EDT$',     lambda _: TODAY),
    (r'.*_DT$',      lambda _: TODAY),
    (r'.*_AMT$',     lambda _: '1000000'),
    (r'.*_RATE$',    lambda _: '5'),
    (r'.*_YN$',      lambda _: 'Y'),
    (r'.*_CD$',      lambda _: '01'),
    (r'.*_TP$',      lambda _: '602-001'),
    (r'.*_CLS$',     lambda _: '604-001'),
    (r'.*_NM$',      lambda _: '테스트명'),
    (r'.*_SBJ$',     lambda _: '테스트구매요청건'),
    (r'.*_RMK$',     lambda _: '테스트비고'),
    (r'FRGN_CLS',    lambda _: '604-001'),
    (r'PUR_TP',      lambda _: '602-001'),
    (r'VAT_TP',      lambda _: '1'),
    (r'ADAMT_FG',    lambda _: '1'),
]


def _get_field_value(col: str) -> str:
    """
    필드명에 대한 테스트 값을 반환한다.
    1순위: md에서 로드된 field_values 딕셔너리 (정확한 필드명 일치)
    2순위: FIELD_VALUE_MAP 패턴 매칭 (기본값)
    """
    # 1순위: md의 필드 테스트 데이터 표에서 정확한 필드명 일치
    md_fields = _RULES.get('field_values', {})
    if col in md_fields:
        return md_fields[col]

    # 2순위: 패턴 매칭
    for pattern, fn in FIELD_VALUE_MAP:
        if re.match(pattern, col):
            return fn(col)
    return ''


# ══════════════════════════════════════════════════════════════════════════════
# 공통코드 DB 런타임 조회
#   - generate_scenarios() 진입 시 _init_comm_code_db() 로 Oracle 연결
#   - 콤보 code_id 발견 시 _get_comm_first(code_id) 로 첫 번째 COMM_CD 반환
#   - 접속 실패 시 COMM_CODE_OVERRIDE 또는 FIELD_VALUE_MAP 으로 폴백
#
# 수동 오버라이드 (DB 값이 아닌 특정 값을 강제 지정할 때):
# COMM_CODE_OVERRIDE = {'602': '602-001', '604': '604-001'}
# ══════════════════════════════════════════════════════════════════════════════
COMM_CODE_OVERRIDE: dict = {}   # {code_id: COMM_CD}  — 필요 시 수동 지정

_DB_CONN       = None           # oracledb / cx_Oracle connection
_DB_AVAILABLE  = None           # None=미시도, True=연결됨, False=실패
_comm_code_cache: dict = {}     # code_id → [(COMM_CD, COMM_NM), ...]
_role_cache: dict    = {}       # pgm_id(대문자) → 'role1, role2' or ''


def _load_db_props(project_root: Path) -> dict:
    """etc/config/tomcat/tomcat.properties 파싱 → {db.host, db.port, db.name, db.user, db.password}"""
    p = project_root / 'etc/config/tomcat/tomcat.properties'
    if not p.exists():
        return {}
    cfg: dict = {}
    for line in p.read_text(encoding='utf-8', errors='replace').splitlines():
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            cfg[k.strip()] = v.strip()
    return cfg


def _init_comm_code_db(project_root: Path) -> None:
    """Oracle 연결 초기화 (generate_scenarios 진입 시 1회 호출)"""
    global _DB_CONN, _DB_AVAILABLE
    if _DB_AVAILABLE is not None:
        return

    cfg  = _load_db_props(project_root)
    host = cfg.get('db.host', '')
    port_str = cfg.get('db.port', '1521')
    sid  = cfg.get('db.name', '')
    user = cfg.get('db.user', '')
    pwd  = cfg.get('db.password', '')

    if not (host and sid and user and pwd):
        print('   [공통코드 DB] 접속 정보 없음 → FIELD_VALUE_MAP 사용')
        _DB_AVAILABLE = False
        return

    port = int(port_str)

    # oracledb (thin 모드, Oracle Instant Client 불필요) 우선 시도
    try:
        import oracledb
        oracledb.init_oracle_client()   # thick 모드가 필요할 경우 경로 지정
    except Exception:
        pass
    try:
        import oracledb
        # SID 기반 DSN — JDBC URL의 ':SID' 형식에 대응
        _sid_dsn = f'(DESCRIPTION=(ADDRESS=(PROTOCOL=TCP)(HOST={host})(PORT={port}))(CONNECT_DATA=(SID={sid})))'
        _DB_CONN      = oracledb.connect(user=user, password=pwd, dsn=_sid_dsn)
        _DB_AVAILABLE = True
        print(f'   [공통코드 DB] 연결 성공 (oracledb): {host}:{port}:{sid}')
        return
    except Exception as e:
        print(f'   [공통코드 DB] oracledb 실패: {e}')
        pass

    # cx_Oracle 폴백
    try:
        import cx_Oracle
        _DB_CONN      = cx_Oracle.connect(user, pwd,
                                           cx_Oracle.makedsn(host, port, sid=sid))
        _DB_AVAILABLE = True
        print(f'   [공통코드 DB] 연결 성공 (cx_Oracle): {host}:{port}/{sid}')
    except Exception as e:
        print(f'   [공통코드 DB] 연결 실패: {e} → FIELD_VALUE_MAP 사용')
        _DB_AVAILABLE = False


def _fetch_comm_codes(code_id: str) -> list:
    """COM_STD_MGT에서 UPP_COMM_CD=code_id 인 코드 목록 조회"""
    if not _DB_CONN:
        return []
    try:
        cur = _DB_CONN.cursor()
        cur.execute("""
            SELECT COMM_CD, COMM_NM
              FROM COM_STD_MGT
             WHERE UPP_COMM_CD = :1
               AND USE_YN = 'Y'
             ORDER BY SORT_NO, COMM_CD
        """, [code_id])
        rows = [(str(r[0]), str(r[1])) for r in cur.fetchall()]
        cur.close()
        return rows
    except Exception as e:
        print(f'   [공통코드 DB] 조회 실패 ({code_id}): {e}')
        return []


def _fetch_role_nm(pgm_id: str) -> str:
    """SYS_PGM_ROLE에서 PGM_ID 기준 ROLE_NM 목록 조회 → 쉼표 결합 문자열 반환"""
    if not _DB_CONN or not pgm_id:
        return ''
    key = pgm_id.upper()
    if key in _role_cache:
        return _role_cache[key]
    try:
        cur = _DB_CONN.cursor()
        cur.execute(
            """SELECT R.ROLE_NM
                 FROM SYS_PGM_USER P
                 JOIN SYS_ROLE_MGT R ON R.ROLE_ID = P.USER_ID
                WHERE P.PGM_ID = :1
                ORDER BY R.ROLE_NM""",
            [key]
        )
        rows = cur.fetchall()
        cur.close()
        result = ', '.join(str(r[0]) for r in rows if r[0])
        print(f'   [역할조회] {key} → {result or "(없음)"}')
        _role_cache[key] = result
        return result
    except Exception as e:
        print(f'   [역할조회] {key} 오류: {e}')
        _role_cache[key] = ''
        return ''


def _get_comm_first(code_id: str) -> str:
    """code_id 그룹의 첫 번째 COMM_CD 반환 (캐시 → DB → 빈 문자열)"""
    if not code_id:
        return ''
    # 수동 오버라이드 우선
    if code_id in COMM_CODE_OVERRIDE:
        return COMM_CODE_OVERRIDE[code_id]
    # 캐시 확인
    if code_id not in _comm_code_cache:
        _comm_code_cache[code_id] = _fetch_comm_codes(code_id)
    rows = _comm_code_cache[code_id]
    return rows[0][0] if rows else ''

# ── XFDL XML 속성값 추출 헬퍼 ───────────────────────────────────────────────
def _xattr(attrs_str: str, name: str) -> str:
    m = re.search(r'\b' + re.escape(name) + r'\s*=\s*["\']([^"\']*)["\']', attrs_str)
    return m.group(1) if m else ''


# ── XFDL 파일 탐색 ──────────────────────────────────────────────────────────
def _find_xfdl(nxui_pur_dir: Path, screen_no: str, pgm_id: str) -> 'Path | None':
    """화면번호/pgm_id로 pur 디렉토리 내 XFDL 파일 탐색"""
    if not nxui_pur_dir or not nxui_pur_dir.exists():
        return None
    candidates = [
        f'pur_{screen_no}M.xfdl', f'pur_{screen_no}m.xfdl',
        f'pur_{screen_no}.xfdl',  f'{pgm_id.lower()}.xfdl',
    ]
    for name in candidates:
        p = nxui_pur_dir / name
        if p.exists():
            return p
    pat = screen_no.lower()
    for f in nxui_pur_dir.glob('*.xfdl'):
        if pat in f.name.lower():
            return f
    return None


# ── XFDL 컴포넌트 파서 ──────────────────────────────────────────────────────
def parse_xfdl_components(xfdl_path: 'Path') -> dict:
    """
    XFDL 파일에서 bind 컬럼명 → 컴포넌트 타입 정보 추출.
    반환: {col_id: {'type': combo|date|number|checkbox|text,
                    'code_id': '604',          # combo+GETCOMCODELIST
                    'inner_vals': [('cd','nm')],# combo+INNERDATASET
                    'true_val': 'Y'}}           # checkbox
    """
    try:
        src = xfdl_path.read_text(encoding='utf-8', errors='replace')
    except Exception:
        return {}

    col_types: dict = {}

    # ① Script → commCd 변수 누적 or gfn_getCommCodeList 리터럴 → ds_codeXXX 매핑
    ds_to_code: dict = {}   # 'ds_code604' → '604'
    raw_strs: list = []
    # commCd = "..."  /  commCd += "..."  /  commCd = commCd + "..."
    for m in re.finditer(r'commCd\s*(?:[+]?=|=\s*commCd\s*[+])\s*["\']([^"\']+)["\']', src):
        raw_strs.append(m.group(1))
    for m in re.finditer(r'gfn_getCommCodeList\s*\(\s*["\']([^"\']+)["\']', src):
        raw_strs.append(m.group(1))
    for s in raw_strs:
        for part in s.split(','):
            raw = part.split(':')[0].strip().lstrip(',').strip()
            cid = raw.replace('-', '').replace(' ', '')
            if cid and re.match(r'^[A-Z0-9]+$', cid, re.IGNORECASE):
                ds_to_code[f'ds_code{cid}']          = cid
                ds_to_code[f'ds_code{cid.zfill(6)}'] = cid  # zero-padded 6자리 변형
                ds_to_code[f'ds_code{raw}']           = cid  # 원본 형태도 등록

    # ② Grid Cell 파싱 (text="bind:COL" + edittype/displaytype)
    for m in re.finditer(r'<Cell\b([^>]*?)/?>', src, re.DOTALL):
        attrs = m.group(1)
        bm = re.search(r'\btext\s*=\s*["\']bind:(\w+)["\']', attrs)
        if not bm:
            continue
        col = bm.group(1)
        if col.startswith('SCH_'):
            continue

        edittype    = _xattr(attrs, 'edittype').lower()
        displaytype = _xattr(attrs, 'displaytype').lower()
        comboset    = _xattr(attrs, 'combodataset')
        true_val    = _xattr(attrs, 'checkboxtruevalue') or 'Y'
        maskfmt     = _xattr(attrs, 'maskeditformat')
        is_mask     = edittype in ('mask', 'number') or bool(maskfmt)
        has_cal     = (('calendardateformat' in attrs or 'dateformat' in attrs) and not is_mask)

        if edittype == 'combo' or 'combo' in displaytype:
            cs_key  = comboset.lstrip('@')
            cs_base = re.sub(r'_\d+$', '', cs_key)
            code_id = ds_to_code.get(cs_key, ds_to_code.get(cs_base, ''))
            if col not in col_types:
                col_types[col] = {'type': 'combo', 'code_id': code_id, 'inner_vals': None}
        elif edittype == 'checkbox' or 'checkbox' in displaytype:
            col_types[col] = {'type': 'checkbox', 'true_val': true_val}
        elif is_mask:
            col_types[col] = {'type': 'number'}
        elif 'date' in displaytype or has_cal:
            col_types[col] = {'type': 'date'}
        elif col not in col_types:
            col_types[col] = {'type': 'text'}

    # ③ 독립 Combo 컴포넌트 파싱 (Div 내)
    #    <Combo id="COL" innerdataset="ds_code604" .../>
    #    또는 INNERDATASET 인라인 정의
    for m in re.finditer(r'<Combo\b([^>]*?)>(.*?)</Combo>|<Combo\b([^>]*?)/>', src, re.DOTALL):
        attrs     = m.group(1) or m.group(3) or ''
        inner_xml = m.group(2) or ''
        comp_id   = _xattr(attrs, 'id')
        if not comp_id or comp_id.startswith('SCH_') or comp_id in col_types:
            continue

        inner_ds      = _xattr(attrs, 'innerdataset').lstrip('@')
        inner_ds_base = re.sub(r'_\d+$', '', inner_ds)
        code_id       = ds_to_code.get(inner_ds, ds_to_code.get(inner_ds_base, ''))

        # INNERDATASET 인라인 Row 추출
        inner_vals = None
        if inner_xml:
            rows = re.findall(
                r'<Row>.*?<Col[^>]*id=["\']codecolumn["\'][^>]*>([^<]*)</Col>'
                r'.*?<Col[^>]*id=["\']datacolumn["\'][^>]*>([^<]*)</Col>',
                inner_xml, re.DOTALL)
            if rows:
                inner_vals = [(cd.strip(), nm.strip()) for cd, nm in rows if cd.strip()]

        col_types[comp_id] = {'type': 'combo', 'code_id': code_id, 'inner_vals': inner_vals}

    # ④ Calendar 컴포넌트 파싱
    for m in re.finditer(r'<Calendar\b([^>]*?)/?>', src, re.DOTALL):
        attrs   = m.group(1)
        comp_id = _xattr(attrs, 'id')
        if comp_id and not comp_id.startswith('SCH_') and comp_id not in col_types:
            col_types[comp_id] = {'type': 'date'}

    return col_types


# ── 컴포넌트 타입 → 테스트 입력값 결정 ───────────────────────────────────────
def _value_for_comp(col: str, info: dict) -> str:
    t = info.get('type', 'text')

    if t == 'combo':
        # 1순위: INNERDATASET 첫 번째 실제 값
        inner = info.get('inner_vals')
        if inner:
            return inner[0][0]
        # 2순위: DB 런타임 조회 (캐시)
        first = _get_comm_first(info.get('code_id', ''))
        if first:
            return first
        # 3순위: FIELD_VALUE_MAP 패턴 폴백
        return _get_field_value(col)

    if t == 'date':
        # _SDT/_FRM/시작일류 → 연초, 나머지 → 오늘
        return YEAR_START if re.search(r'_SDT$|_FRM$|_BGN$|_STR$', col) else TODAY

    if t == 'number':
        return _get_field_value(col) or '1000000'

    if t == 'checkbox':
        return info.get('true_val', 'Y')

    return _get_field_value(col)   # text


def build_test_data_from_cols(cols: list, crud: str,
                               col_types: 'dict | None' = None) -> str:
    """INSERT 컬럼 목록 → 테스트 데이터 문자열"""
    if not cols:
        return f'RQST_SDT={YEAR_START}; RQST_EDT={TODAY}' if crud == 'SELECT' else ''
    skip = {'INS_ID', 'INS_DE', 'INS_IP', 'UPT_ID', 'UPT_DE', 'UPT_IP',
            'SESS_USER_ID', 'SESS_IP', 'SYSDATE'}
    parts = []
    for col in cols:
        if col in skip:
            continue
        info = (col_types or {}).get(col)
        val  = _value_for_comp(col, info) if info else _get_field_value(col)
        if val:
            parts.append(f'{col}={val}')
    return '; '.join(parts) if parts else ''


def build_test_data_for_screen(screen_no: str, crud: str, insert_cols_map: dict,
                                nxui_pur_dir: 'Path | None' = None,
                                pgm_id: str = '') -> str:
    """
    화면 번호 기반 테스트 데이터 생성.
    우선순위: XFDL 컴포넌트 타입 > md 필드값 > SCREEN_FLOW 필수필드 > SQLMAP 컬럼
    XFDL 파싱은 CRUD=SAVE/INSERT/UPDATE 인 경우만 수행.
    """
    # XFDL 컴포넌트 타입 로드 (SAVE 계열만)
    col_types: dict = {}
    if crud in ('SAVE', 'INSERT', 'UPDATE') and nxui_pur_dir:
        xfdl_path = _find_xfdl(nxui_pur_dir, screen_no, pgm_id)
        if xfdl_path:
            col_types = parse_xfdl_components(xfdl_path)

    flow = SCREEN_FLOW.get(screen_no)
    if flow and crud in ('SAVE', 'INSERT', 'UPDATE'):
        req = flow.get('required_fields', []) + flow.get('required_amount_fields', [])
        if req:
            parts = []
            for f in req:
                info = col_types.get(f)
                val  = _value_for_comp(f, info) if info else _get_field_value(f)
                if val:
                    parts.append(f'{f}={val}')
            if parts:
                return '; '.join(parts)

    # SQLMAP INSERT 컬럼 fallback (키 형식: "{screen_no}::{qid}", SAVE 계열만 적용)
    if crud in ('SAVE', 'INSERT', 'UPDATE'):
        for k, cols in insert_cols_map.items():
            file_sno, _, _ = k.partition('::')
            if file_sno == screen_no:
                return build_test_data_from_cols(cols, crud, col_types or None)

    # XFDL col_types fallback — SCREEN_FLOW·SQLMAP 미등록 화면도 입력값 생성
    if col_types and crud in ('SAVE', 'INSERT', 'UPDATE'):
        _UI_SKIP = {'isChecked', 'tmHeader', 'SCH_'}
        meaningful = [col for col in col_types
                      if not any(col == s or col.startswith('SCH_') for s in _UI_SKIP)]
        if meaningful:
            return build_test_data_from_cols(meaningful, crud, col_types)

    if crud == 'SELECT':
        return f'RQST_SDT={YEAR_START}; RQST_EDT={TODAY}'
    return ''

# ══════════════════════════════════════════════════════════════════════════════
# 5. 시나리오 빌더
# ══════════════════════════════════════════════════════════════════════════════

def make_pre_condition(crud: str, screen_no: str, flow_info: dict = None) -> str:
    flow = flow_info or SCREEN_FLOW.get(screen_no, {})
    steps = ['로그인 및 구매 처리 권한 확인']
    if flow.get('pre_screen'):
        steps.append(
            f'PUR_{flow["pre_screen"]} {flow.get("name","이전화면")} 완료'
            f' ({flow.get("pre_stat","결재완료")} 상태) 확인')
    if crud in ('SELECT',):
        steps.append('조회 화면 진입')
        steps.append('조회 조건 입력')
        steps.append('조회 버튼 클릭')
        steps.append('조회 결과 목록 확인')
    elif crud in ('SAVE', 'INSERT', 'UPDATE'):
        steps.append('입력 화면 진입')
        steps.append('필수 항목 입력 (입력값 컬럼 참조)')
        steps.append('저장 버튼 클릭')
        steps.append('저장 결과 확인 (재조회 또는 완료 메시지)')
    elif crud == 'DELETE':
        steps.append('삭제 대상 건 선택 (APV_STAT_CD=임시저장 상태)')
        steps.append('삭제 버튼 클릭')
        steps.append('삭제 결과 확인 (목록에서 제거 여부)')
    elif crud == 'APPROVAL':
        steps.append('결재선 설정 완료 확인')
        steps.append('결재 요청 버튼 클릭')
        steps.append('결재 상태 변경 확인')
    elif crud in ('CHECK',):
        steps.append('확인 대상 건 선택')
        steps.append('확인 처리 버튼 클릭')
        steps.append('처리 결과 확인')
    else:
        steps.append('해당 화면 진입')
        steps.append('처리 버튼 클릭')
        steps.append('처리 결과 확인')
    return '\n'.join(f'{i+1}. {s}' for i, s in enumerate(steps))


def make_expected(crud: str, screen_desc: str) -> str:
    """
    CRUD 유형별 예상결과 문구를 반환한다.
    1순위: md의 섹션 6 (crud_expected)에서 로드된 문구
    2순위: 내부 기본값 (_CRUD_EXPECTED_DEFAULT)
    {화면명} / {screen_desc} 플레이스홀더는 실제 화면명으로 치환된다.
    """
    # 1순위: md 로드 문구
    md_expected = _RULES.get('crud_expected', {})
    template = md_expected.get(crud) or _CRUD_EXPECTED_DEFAULT.get(crud)
    if not template:
        return f'{screen_desc} 처리가 정상 완료되고 result=Y가 반환된다.'

    # {화면명} / {screen_desc} 플레이스홀더 치환
    return (template
            .replace('{화면명}', screen_desc)
            .replace('{screen_desc}', screen_desc)
            .replace('`{화면명}`', screen_desc))


def build_scenario(no, test_type, sid, screen_no, screen_name, category,
                   test_name, pgm_id, crud, pre, expected, test_data, remark='',
                   description='') -> dict:
    return dict(no=no, testType=test_type, scenarioId=sid,
                screenNo=screen_no, screenName=screen_name, category=category,
                testName=test_name, description=description or test_name,
                pgmId=pgm_id, httpMethod='POST', crudType=crud,
                preCondition=pre, testData=test_data,
                expectedResult=expected, testResult='', remark=remark,
                # 중분류/소분류/메뉴명: Python 생성 시에는 빈 값, Java DB 조회로 채워짐
                midCategory='', subCategory='', menuName='')

# ══════════════════════════════════════════════════════════════════════════════
# 6. 메인 생성 엔진
# ══════════════════════════════════════════════════════════════════════════════

def generate_scenarios(project_root: Path, source_filter: set = None) -> list:
    ctrl_dir      = project_root / 'src/main/java/cres/mis/pur/web'
    svc_dir       = project_root / 'src/main/java/cres/mis/pur/service/impl'
    sqlmap_dir    = project_root / 'src/main/resources/cres/sqlmap/mis/pur'
    nxui_pur_dir  = project_root / f'src/main/nxui/{os.environ["APP_CONTEXT_PATH"]}/mis/pur'

    if not ctrl_dir.exists():
        print(f'[ERROR] Controller 디렉토리 없음: {ctrl_dir}')
        return []

    # ── 0. 공통코드 DB 연결 (콤보 입력값 조회용) ────────────────────────────
    _init_comm_code_db(project_root)

    # ── 1. SQLMAP 전체 인덱스 ────────────────────────────────────────────────
    print('[1/4] SQLMAP XML 파싱...')
    sqlmap_index    = {}
    insert_cols_map = {}  # "{screen_no}::{qid}" → [cols]
    if sqlmap_dir.exists():
        for xf in sorted(sqlmap_dir.glob('*.xml')):
            r      = parse_sqlmap(xf)
            sqlmap_index.update(r.get('crud', {}))
            # 파일명에서 화면 번호 추출 → 키를 화면별로 스코핑 (Pur_5110.xml → "5110")
            file_sno = extract_screen_no(xf.stem)
            for qid, cols in r.get('insert_cols', {}).items():
                scoped_key = f'{file_sno}::{qid}' if file_sno else qid
                insert_cols_map[scoped_key] = cols
    print(f'      → {len(sqlmap_index)}개 쿼리, {len(insert_cols_map)}개 INSERT 컬럼 세트')

    # ── 2. ServiceImpl 파싱 (예외 패턴) ────────────────────────────────────
    print('[2/4] ServiceImpl 예외 패턴 파싱...')
    svc_map = {}   # screen_no → service info
    if svc_dir.exists():
        for jf in sorted(svc_dir.glob('*ServiceImpl.java')):
            sno = extract_screen_no(jf.name)
            parsed = parse_service_impl(jf)
            if sno and parsed:
                svc_map[sno] = parsed
    total_exc = sum(len(s.get('global_exceptions', [])) +
                    sum(len(m['exceptions']) for m in s.get('methods', []))
                    for s in svc_map.values())
    print(f'      → {len(svc_map)}개 ServiceImpl, {total_exc}개 예외 패턴')

    # ── 3. Controller 파싱 ──────────────────────────────────────────────────
    print('[3/4] Controller 파싱...')
    if source_filter:
        print(f'      → 소스 필터 적용: {sorted(source_filter)}')
    controllers = []
    for jf in sorted(ctrl_dir.glob('*Controller.java')):
        if source_filter:
            sno = extract_screen_no(jf.name)
            if sno not in source_filter:
                continue
        parsed = parse_controller(jf)
        if parsed and parsed.get('endpoints'):
            controllers.append(parsed)
    print(f'      → {len(controllers)}개 Controller, '
          f'{sum(len(c["endpoints"]) for c in controllers)}개 엔드포인트')

    # ── 4. 시나리오 생성 ────────────────────────────────────────────────────
    print('[4/4] 시나리오 생성...')
    scenarios = []
    no = 1

    for ctrl in controllers:
        screen_no   = ctrl['screen_no']
        pgm_id      = ctrl.get('pgm_id', f'PUR_{screen_no}M')
        cat_info    = SCREEN_CATEGORY.get(screen_no, ('구매', f'PUR_{screen_no} 화면'))
        category    = cat_info[0]
        screen_desc = cat_info[1]
        screen_name = f'PUR_{screen_no}' if screen_no else ctrl['filename'].replace('Controller.java', '')
        flow_info   = SCREEN_FLOW.get(screen_no, {})
        svc_info    = svc_map.get(screen_no, {})

        has_save    = any(classify_crud(e['url'], e['method'], sqlmap_index)
                         in ('SAVE','INSERT','UPDATE') for e in ctrl['endpoints'])
        has_delete  = any(classify_crud(e['url'], e['method'], sqlmap_index)
                         == 'DELETE' for e in ctrl['endpoints'])
        has_approx  = any('app' in e['method'].lower() or 'app' in e['url'].lower()
                          for e in ctrl['endpoints'])
        has_budgchk = any('budg' in e['url'].lower() or 'checkbudg' in e['method'].lower()
                          for e in ctrl['endpoints'])
        has_conc    = any('uptChk' in e['method'] or 'uptchk' in e['url'].lower()
                          for e in ctrl['endpoints'])

        # ── [A] 단위 시나리오: 엔드포인트별 1개 ─────────────────────────────
        for ep in ctrl['endpoints']:
            crud = classify_crud(ep['url'], ep['method'], sqlmap_index)
            pre  = make_pre_condition(crud, screen_no, flow_info)
            exp  = make_expected(crud, screen_desc)
            data = build_test_data_for_screen(screen_no, crud, insert_cols_map, nxui_pur_dir, pgm_id)
            ep_label = ep['url'].rsplit('/', 1)[-1].replace('.do', '')
            sid = f'UT_{pgm_id}'
            scenarios.append(build_scenario(
                no, '단위', sid, screen_no, screen_name, category,
                f'[{category}] {screen_desc} — {ep["comment"] or ep_label} ({crud})',
                pgm_id, crud, pre, exp, data, ep.get('comment', ''),
                description=ep.get('comment') or f'{screen_desc} {ep_label} {crud} 검증'))
            no += 1

        # ── [B] 통합 시나리오: CRUD 흐름 ────────────────────────────────────
        if has_save:
            pre  = make_pre_condition('SAVE', screen_no, flow_info)
            data = build_test_data_for_screen(screen_no, 'SAVE', insert_cols_map, nxui_pur_dir, pgm_id)
            exp  = (f'{screen_desc} 저장 후 재조회 시 데이터가 반영되어 있어야 한다. '
                    f'관련 화면(예산현황, 집행이력)에서도 변경 내용이 확인된다.')
            scenarios.append(build_scenario(
                no, '통합', f'IT_{pgm_id}', screen_no, screen_name, category,
                f'[{category}] {screen_desc} — 저장·재조회 통합 흐름',
                pgm_id, 'SAVE', pre, exp, data, '조회→저장→재조회 3단계 통합 검증',
                description=f'{screen_desc} 저장 후 재조회 시 데이터가 정상 반영되는지 검증한다.'))
            no += 1

        if has_delete:
            pre  = make_pre_condition('DELETE', screen_no, flow_info)
            data = f'삭제 대상 {flow_info.get("key","RQST_NO")} (임시저장 상태 건)'
            exp  = (f'삭제 처리 후 목록에서 제거되고, '
                    f'{flow_info.get("key","RQST_NO")}로 재조회 시 결과가 없어야 한다.')
            scenarios.append(build_scenario(
                no, '통합', f'IT_{pgm_id}', screen_no, screen_name, category,
                f'[{category}] {screen_desc} — 삭제·재조회 통합 흐름',
                pgm_id, 'DELETE', pre, exp, data, '삭제→재조회 확인',
                description=f'{screen_desc} 삭제 처리 후 목록에서 제거되는지 검증한다.'))
            no += 1

        # ── [C] 비정상 시나리오: 상태 제약 ──────────────────────────────────
        # C-1. 결재 진행 중 수정 불가
        if has_save or has_delete:
            for st_name, st in STATE_MACHINE.items():
                if not st['blocked']:
                    continue
                sid = f'IT_{pgm_id}'
                scenarios.append(build_scenario(
                    no, '통합', sid, screen_no, screen_name, category,
                    f'[{category}·상태제약] {screen_desc} — {st_name} 상태에서 수정/삭제 시도',
                    pgm_id, '비정상',
                    f'{st["test_data"]} 상태 건 선택 후 {", ".join(st["blocked"][:2])} 시도',
                    f'{st["blocked"][:2]} 작업이 차단되고 오류 메시지가 반환된다.',
                    st['test_data'],
                    f'차단 작업: {", ".join(st["blocked"])}',
                    description=f'{st_name} 상태에서 {", ".join(st["blocked"][:2])} 시도 시 차단 여부를 검증한다.'))
                no += 1
                break   # 대표 1개만

        # C-2. 예산 잔액 부족
        if has_budgchk:
            scenarios.append(build_scenario(
                no, '통합', f'IT_{pgm_id}', screen_no, screen_name, category,
                f'[{category}·예산검증] {screen_desc} — 예산 잔액 부족 시 저장 차단',
                pgm_id, '비정상',
                '예산 과목 잔액을 초과하는 금액 입력 (TOT_PRES_AMT > 예산 잔액)',
                'CHK_OK=N과 부족 금액이 포함된 CHK_MSG가 반환되고 저장이 차단된다.',
                f'TOT_PRES_AMT=99999999; BUDG_ITSR_CD=해당과목코드',
                '예산 잔액 초과 케이스',
                description='예산 잔액 부족 시 저장이 차단되고 오류 메시지가 반환되는지 검증한다.'))
            no += 1
            # 정상 케이스
            scenarios.append(build_scenario(
                no, '통합', f'IT_{pgm_id}', screen_no, screen_name, category,
                f'[{category}·예산검증] {screen_desc} — 예산 잔액 충분 시 저장 허용',
                pgm_id, 'CHECK',
                '예산 과목 잔액이 충분한 금액 입력 (TOT_PRES_AMT ≤ 예산 잔액)',
                'CHK_OK=Y가 반환되고 저장이 정상 완료된다.',
                f'TOT_PRES_AMT=100000; BUDG_ITSR_CD=잔액충분한과목코드',
                '예산 잔액 충분 케이스',
                description='예산 잔액이 충분할 때 저장이 정상 허용되는지 검증한다.'))
            no += 1

        # C-3. 동시성 체크
        if has_conc:
            scenarios.append(build_scenario(
                no, '통합', f'IT_{pgm_id}', screen_no, screen_name, category,
                f'[{category}·동시성] {screen_desc} — 동시 수정 충돌 시 저장 차단',
                pgm_id, '비정상',
                '동일 건에 대해 두 세션에서 동시에 수정 후 저장 시도',
                '나중에 저장 시도한 세션에서 result=N이 반환되고 저장이 차단된다.',
                '세션1에서 먼저 저장 완료 후 세션2에서 동일 건 저장 시도',
                '동시성 충돌 케이스',
                description='동시 수정 시 나중에 저장한 세션이 차단되는지 검증한다.'))
            no += 1

        # ── [D] ServiceImpl 예외 기반 시나리오 ────────────────────────────
        svc_methods = svc_info.get('methods', [])
        for method in svc_methods:
            for exc in method.get('exceptions', []):
                sid = f'IT_{pgm_id}'
                scenarios.append(build_scenario(
                    no, '통합', sid, screen_no, screen_name, category,
                    f'[{category}·서비스예외] {screen_desc} — {method["name"]}: {exc["msg"][:40]}',
                    pgm_id, '비정상',
                    f'{exc["type"]} 발생 조건: {exc["msg"][:80]}',
                    f'"{exc["msg"]}" 예외 메시지가 반환되고 처리가 중단된다.',
                    '예외 발생 조건에 맞는 데이터 입력 필요',
                    f'ServiceImpl.{method["name"]} 예외 케이스',
                    description=f'{method["name"]} 서비스 예외 발생 시 "{exc["msg"][:40]}" 오류가 반환되는지 검증한다.'))
                no += 1
            for cond in method.get('conditions', []):
                if len(cond.get('condition', '')) < 5:
                    continue
                sid = f'IT_{pgm_id}'
                scenarios.append(build_scenario(
                    no, '통합', sid, screen_no, screen_name, category,
                    f'[{category}·조건검증] {screen_desc} — {method["name"]} 조건 불충족',
                    pgm_id, '비정상',
                    f'조건: {cond["condition"][:80]}',
                    f'{cond["consequence"][:80]}',
                    '조건 불충족 상태 데이터 구성 필요',
                    f'ServiceImpl 조건: {cond["condition"][:40]}',
                    description=f'{method["name"]} 조건 불충족 시 처리가 차단되는지 검증한다. 조건: {cond["condition"][:60]}'))
                no += 1

        # ── [E] 화면 간 선행 조건 시나리오 ──────────────────────────────────
        if flow_info.get('pre_screen'):
            scenarios.append(build_scenario(
                no, '통합', f'IT_{pgm_id}', screen_no, screen_name, category,
                f'[{category}·선행조건] {screen_desc} — 선행 단계 미완료 시 진행 불가',
                pgm_id, '비정상',
                f'선행 화면 PUR_{flow_info["pre_screen"]} 결재 미완료 상태 ({flow_info.get("pre_stat","결재완료")} 전)',
                f'선행 화면({flow_info["pre_screen"]})의 처리가 완료되지 않은 경우 '
                f'{screen_desc}으로의 진행이 차단된다.',
                f'선행 {flow_info.get("pre_key","RQST_NO")} 미입력 또는 결재진행 상태',
                f'선행조건: PUR_{flow_info["pre_screen"]} {flow_info.get("pre_stat","")} 필요',
                description=f'PUR_{flow_info["pre_screen"]} 미완료 상태에서 {screen_desc} 진행 시 차단되는지 검증한다.'))
            no += 1

    # ── [F] E2E 통합 흐름 시나리오 ──────────────────────────────────────────
    e2e_flows = [
        {
            'name': '국내 구매 전체 흐름 (구매요청 → 정산완료)',
            'steps': [
                '1. PUR_0010 구매요청 등록 → RQST_NO 발급',
                '2. PUR_0010 결재 요청 → APV_STAT_CD=020',
                '3. (결재 완료) → APV_STAT_CD=040, PUR_STEP=601-010',
                '4. PUR_0150 구매계약 목록 조회 → 해당 건 확인',
                '5. PUR_0153 구매계약 상세 등록 → 예산체크(CHK_OK=Y) 후 저장',
                '6. PUR_0200 발주 대상 확인 → 발주 처리',
                '7. PUR_0300 납품 등록 → DLVR_NO 발급',
                '8. PUR_0400 대가지급 등록 → PAY_NO 발급',
                '9. PUR_0500 정산 완료 → PUR_STEP=601-080',
            ],
            'expected': '전체 구매 프로세스가 단계별로 진행되고 각 단계의 PUR_STEP이 정상 전이된다. '
                        '마지막 정산완료 후 PUR_0610 예산현황에 금액이 반영된다.',
        },
        {
            'name': '구매계약 취소 흐름 (계약 후 취소)',
            'steps': [
                '1. PUR_0153 구매계약 상세 등록 완료 상태 준비',
                '2. 계약취소 요청 (setPurDelCont) → PUR_STEP=601-091',
                '3. 원인행위 삭제 확인 (BDG_CTRL 삭제 확인)',
                '4. PUR_0150 목록에서 취소 상태로 표시 확인',
            ],
            'expected': 'PUR_STEP이 601-091(계약취소)로 변경되고, 원인행위(예산 통제)가 삭제되어 '
                        '예산 잔액이 복원된다.',
        },
        {
            'name': '정산취소 흐름 (자산 미등록 조건)',
            'steps': [
                '1. PUR_0500 정산완료(PUR_STEP=601-080) 상태 준비',
                '2. 자산 미등록 확인 (chkAssRegCnt=0)',
                '3. 정산취소 (setEndSaveRtn) → PUR_STEP=601-070',
            ],
            'expected': 'chkAssRegCnt=0 조건에서 정산이 취소되고 PUR_STEP=601-070으로 변경된다.',
        },
        {
            'name': '정산취소 불가 흐름 (자산 등록된 경우)',
            'steps': [
                '1. PUR_0500 정산완료 상태 준비',
                '2. 자산 등록 존재 (chkAssRegCnt > 0)',
                '3. 정산취소 시도 → SystemException 발생',
            ],
            'expected': '"현재 자산이 등록되어 있는 계약 건은 정산취소 할 수 없습니다." 메시지가 반환된다.',
        },
        {
            'name': '해외 구매 전체 흐름',
            'steps': [
                '1. PUR_5110 해외구매 요청 → 환율 적용 확인',
                '2. PUR_5120 해외계약 등록',
                '3. PUR_5130 해외발주',
                '4. PUR_5140 해외납품 등록',
                '5. PUR_5150 해외대가지급',
            ],
            'expected': '해외 구매 프로세스가 순서대로 진행되고 환율(REP_EXCHNG_RATE)이 각 단계에 올바르게 적용된다.',
        },
    ]
    # E2E는 소스 필터 없이 전체 PUR 생성할 때만 추가
    if not source_filter:
        for flow in e2e_flows:
            scenarios.append(build_scenario(
                no, '통합', f'IT_PUR_E2E',
                'E2E', 'PUR_E2E', 'E2E흐름',
                flow['name'],
                'PUR_E2E',
                'E2E',
                '로그인 상태, 각 단계 처리 권한, 테스트 예산 과목 준비',
                flow['expected'],
                '단계별 테스트 데이터 별도 준비 필요',
                ' → '.join(flow['steps'][:4]) + (' ...' if len(flow['steps']) > 4 else ''),
                description=flow['name'] + ' — ' + ' → '.join(flow['steps'][:3]) + (' ...' if len(flow['steps']) > 3 else '')))
            no += 1

    # ── 후처리: menuName이 비어있고 SCREEN_CATEGORY에 실제 등록된 화면만 설정 ──
    # SCREEN_CATEGORY 기본값("PUR_XXXX 화면")은 넣지 않음 → Java DB 조회값이 우선
    for s in scenarios:
        if not s.get('menuName'):
            sno = s.get('screenNo', '')
            if sno in SCREEN_CATEGORY:
                s['menuName'] = SCREEN_CATEGORY[sno][1]

    # ── 후처리: 순번(no)을 scenarioId 그룹 내 1, 2, 3, ... 으로 재번호 ──────────
    # scenarioId가 같은 행끼리 순번을 1부터 카운트
    _no_map: dict = {}
    for s in scenarios:
        sid = s['scenarioId']
        _no_map[sid] = _no_map.get(sid, 0) + 1
        s['no'] = _no_map[sid]

    # ── 역할 조회 (SYS_PGM_ROLE) ───────────────────────────────────────────────
    if _DB_AVAILABLE:
        print('[역할조회] SYS_PGM_ROLE DB 조회...')
        for s in scenarios:
            pgm_id = s.get('pgmId', '')
            if pgm_id and not s.get('roleNm'):
                s['roleNm'] = _fetch_role_nm(pgm_id)
    else:
        print('[역할조회] DB 미연결 — roleNm 빈값 (Java enrichRoles에서 채움)')

    print(f'      → 총 {len(scenarios)}개 시나리오')
    return scenarios

# ══════════════════════════════════════════════════════════════════════════════
# 7. Excel 출력
# ══════════════════════════════════════════════════════════════════════════════

HEADERS = [
    ('No', 5), ('구분', 8), ('시나리오ID', 16), ('화면번호', 10),
    ('화면명', 14), ('카테고리', 12), ('테스트명', 52),
    ('소스명(pgmId)', 18), ('CRUD', 10), ('사전조건', 38), ('테스트데이터', 42),
    ('예상결과', 56), ('테스트결과', 12), ('비고', 30),
]
COLOR = {'단위': 'DBEAFE', '통합': 'DCFCE7'}
CRUD_COLOR = {
    'SELECT': 'E0F2FE', 'SAVE': 'DCFCE7', 'INSERT': 'DCFCE7', 'UPDATE': 'FEF3C7',
    'DELETE': 'FFE4E6', 'CHECK': 'F3E8FF', 'APPROVAL': 'FFF7ED',
    'CANCEL': 'FDE8D0', 'NEGATIVE': 'FEE2E2', 'E2E': 'FEF9C3',
}

def _border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def write_excel(scenarios: list, out_path: Path, template_path: Path = None):
    # 엑터 매핑
    ACTOR_MAP = {'단위': '개발자', '통합': '업무담당자', '비정상': '테스터', 'E2E': '업무담당자'}

    # ── 양식 파일 로드 ────────────────────────────────────────────────────────
    use_template = False
    wb = None
    ws = None

    if template_path and template_path.exists():
        try:
            wb = openpyxl.load_workbook(str(template_path))
            # '시나리오및결과서' 시트 탐색
            target = None
            for sn in wb.sheetnames:
                if '시나리오' in sn:
                    target = sn
                    break
            if target is None:
                target = wb.sheetnames[-1]
            ws = wb[target]
            # 3행 이후 데이터 삭제 (헤더 행 1~2 유지)
            if ws.max_row >= 3:
                ws.delete_rows(3, ws.max_row - 2)
            use_template = True
            print(f'   양식 시트 적용: {target}')
        except Exception as e:
            print(f'   [WARN] 양식 로드 실패 ({e}) → 기본 형식으로 생성')
            wb = None

    # ── 양식 없을 때 기본 형식 생성 ──────────────────────────────────────────
    if wb is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '시나리오및결과서'
        ws.freeze_panes = 'A3'

        # 헤더 행 1: 그룹 타이틀
        # 통일 스키마 — JSP PUR_COLS / Java downloadEditedScenarios 헤더와 동일
        # 통일 스키마 — JSP PUR_COLS / STEP2 테이블 / Java downloadEditedScenarios 헤더와 동일
        TMPL_COLS = ['No','구분','시나리오ID','테스트명','엑터(역할)',
                     'URL','Method','대메뉴(GNB)','중분류','소분류','메뉴명',
                     '시나리오흐름','입력값','예상결과','테스트결과','확인자','판정결과',
                     'PL확인','사유','사용자확인','메뉴경로']
        COL_W     = [5.0, 6.0, 13.0, 28.0, 12.0,
                     22.0, 8.0, 12.0, 12.0, 12.0, 16.0,
                     30.0, 18.0, 30.0, 12.0, 8.0, 10.0,
                     8.0, 16.0, 8.0, 24.0]
        ncols = len(TMPL_COLS)
        # 시나리오 정보: 1열(No) ~ 14열(예상결과) / 테스트결과 그룹: 15열(테스트결과) ~ ncols열(메뉴경로)
        _SCEN_LAST = 14   # '예상결과' 위치 (1-indexed) — TMPL_COLS 변경 시에도 고정
        ws.merge_cells(f'A1:{get_column_letter(_SCEN_LAST)}1')
        ws['A1'].value = f'통합테스트 시나리오  (생성일: {date.today()})'
        ws['A1'].font      = Font(bold=True, size=12, color='FFFFFF')
        ws['A1'].fill      = PatternFill('solid', fgColor='1E3A5F')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        # 테스트결과 그룹 (테스트결과~메뉴경로)
        ws.merge_cells(f'{get_column_letter(_SCEN_LAST + 1)}1:{get_column_letter(ncols)}1')
        ws[f'{get_column_letter(_SCEN_LAST + 1)}1'].value = '테스트결과'
        ws[f'{get_column_letter(_SCEN_LAST + 1)}1'].font      = Font(bold=True, size=10, color='FFFFFF')
        ws[f'{get_column_letter(_SCEN_LAST + 1)}1'].fill      = PatternFill('solid', fgColor='1E3A5F')
        ws[f'{get_column_letter(_SCEN_LAST + 1)}1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24

        # 헤더 행 2
        for ci, (h, w) in enumerate(zip(TMPL_COLS, COL_W), 1):
            cell = ws.cell(2, ci, h)
            cell.font      = Font(bold=True, color='FFFFFF', size=10)
            cell.fill      = PatternFill('solid', fgColor='2563EB')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 22

    # ── 양식 헤더 읽기 → 컬럼 인덱스 동적 매핑 ─────────────────────────────
    # 헤더 행 1(ri=1)과 행 2(ri=2)를 합산: 2행 값 우선, 없으면 1행 값
    def _read_headers(ws):
        h1_row = ws[1]
        h2_row = ws[2] if ws.max_row >= 2 else []
        h2 = {c.column: (c.value or '').strip() for c in h2_row}
        headers = {}   # 헤더명(소문자) → 1-based 컬럼 인덱스
        for c in h1_row:
            ci   = c.column
            name = (c.value or '').strip()
            name = h2.get(ci, '') or name   # 2행 값 우선
            if name:
                headers[name] = ci
                headers[name.lower()] = ci
        return headers

    col_map = _read_headers(ws)

    # ── 소스명(pgmId) 컬럼 강제 보장 ─────────────────────────────────────────
    # 양식(템플릿)에 해당 컬럼이 없으면 숨김 컬럼을 추가해 Java가 pgmId를 읽을 수 있게 한다.
    _PGM_VARIANTS = {'소스명(pgmid)', '소스명', 'pgmid', 'pgm_id', 'sourcename'}
    if not any(k in _PGM_VARIANTS for k in col_map):
        _pgm_col = (max(col_map.values()) if col_map else 0) + 1
        _pgm_key = '소스명(pgmId)'
        col_map[_pgm_key]         = _pgm_col
        col_map[_pgm_key.lower()] = _pgm_col
        ws.cell(2, _pgm_col, _pgm_key)
        ws.column_dimensions[get_column_letter(_pgm_col)].hidden = True
        print(f'   [pgmId 보강] 숨김 컬럼 {get_column_letter(_pgm_col)} 에 소스명(pgmId) 추가')

    # 헤더명 → 시나리오 데이터 getter 매핑
    # 키: 양식에서 사용할 가능한 헤더명들 (한글·영문 모두 등록)
    def _scenario_val(header_name, s, flow):
        h = header_name.lower().strip()
        if   h in ('시나리오id','scenarioid','id'):        return s.get('scenarioId','')
        elif h in ('시나리오명','screenname'):              return s.get('screenName','') or s.get('menuName','')
        elif h in ('설명','description'):                   return s.get('description','') or s.get('testName','')
        elif h in ('테스트명','testname'):                  return s.get('testName','')
        elif h in ('순번','no','번호'):                     return s.get('no','')
        elif h in ('구분','testtype','유형'):               return s.get('testType','')
        elif h in ('엑터(역할)','엑터','actor','역할','rolename','rolnm','rolenm'): return s.get('roleNm','')
        elif h in ('url',):                                 return s.get('url','')
        elif h in ('method','메소드'):                      return s.get('method','')
        elif h in ('대메뉴(gnb)','gnbname','대메뉴','gnb'): return s.get('gnbName','')
        elif h in ('중분류',):                              return s.get('midCategory','') or s.get('groupName','') or s.get('category','')
        elif h in ('소분류','subcategory'):                 return s.get('subCategory','')
        elif h in ('메뉴명','menu','menuname'):             return s.get('menuName','')
        elif h in ('메뉴경로','menupath','경로'):           return s.get('menuPath','')
        elif h in ('category','카테고리'):                  return s.get('category','')
        elif h in ('시나리오흐름','흐름','flow','시나리오 흐름','precondition','사전조건'): return flow
        elif h in ('입력값','inputvalues','입력 값','testdata','테스트데이터'): return s.get('inputValues','') or s.get('testData','')
        elif h in ('예상결과','expectedresult','기대결과'): return s.get('expectedResult','')
        elif h in ('확인자','checker','tester','confirmer'): return s.get('confirmer','')
        elif h in ('판정결과','judgmentresult','판정'):     return s.get('judgmentResult','')
        elif h in ('테스트결과','testresult','result'):     return s.get('testResult','')
        elif h in ('pl확인','plconfirm','pl 확인'):         return s.get('plConfirm','')
        elif h in ('사유','reason'):                         return s.get('reason','')
        elif h in ('사용자확인','userconfirm','사용자 확인'): return s.get('userConfirm','')
        elif h in ('소스명(pgmid)','소스명','pgmid','pgm_id','sourcename'): return s.get('pgmId','')
        elif h in ('crud','crudtype'):                      return s.get('crudType','')
        elif h in ('비고','remark','note'):                 return s.get('remark','')
        elif h in ('화면번호','screenno'):                  return s.get('screenNo','')
        return ''

    # 매핑 결과 출력 (디버깅)
    print(f'   컬럼 매핑: {len(col_map)}개 헤더 인식')

    # ── 데이터 행 작성 ────────────────────────────────────────────────────────
    for ri, s in enumerate(scenarios, 3):
        flow = s.get('preCondition', '')

        bg = COLOR.get(s['testType'], 'FFFFFF')

        # 매핑된 컬럼에만 값을 씀
        written_cols = set()
        for header_name, ci in col_map.items():
            if ci in written_cols:
                continue
            val = _scenario_val(header_name, s, flow)
            if val == '' and not header_name.lower() in ('확인자','판정결과','checker','tester'):
                pass   # 빈 값도 기록
            cell = ws.cell(ri, ci, val)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border    = _border()
            cell.font      = Font(size=9)
            h_lower = header_name.lower()
            if h_lower in ('시나리오id','시나리오명','scenarioid','screenname'):
                cell.fill = PatternFill('solid', fgColor=bg)
            elif h_lower in ('엑터(역할)','엑터','actor','역할'):
                cell.fill = PatternFill('solid', fgColor='F3F4F6')
            else:
                cell.fill = PatternFill('solid', fgColor='FFFFFF' if ri % 2 == 0 else 'F8FAFC')
            written_cols.add(ci)

        ws.row_dimensions[ri].height = 58

    if not use_template:
        ws.auto_filter.ref = f'A2:{get_column_letter(max(col_map.values()) if col_map else 12)}2'

    _safe_save(wb, out_path)
    print(f'   시나리오및결과서: {len(scenarios)}건')


def _safe_save(wb, out_path: Path):
    """
    openpyxl Workbook 저장.
    파일이 열려있거나 권한 오류 시 대체 경로로 재시도한다.
    대체 경로 순서:
      1. 원래 경로 (out_path)
      2. 스크립트와 같은 폴더 (copilot_info/)
      3. 사용자 홈 디렉토리 (~/Desktop 또는 ~/)
    """
    candidates = [
        out_path,
        Path(__file__).parent / out_path.name,          # copilot_info/
        Path.home() / 'Desktop' / out_path.name,        # 바탕화면
        Path.home() / out_path.name,                    # 홈 디렉토리
    ]
    for path in candidates:
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(path))
            print(f'\n✅ Excel 저장 완료: {path}')
            return
        except PermissionError:
            print(f'[WARN] 저장 실패 (파일이 열려있거나 권한 없음): {path}')
            if path == candidates[-1]:
                print('[ERROR] 모든 경로 저장 실패.')
                print('        Excel에서 이전 파일을 닫은 후 다시 실행하세요.')
                raise
        except Exception as e:
            print(f'[WARN] 저장 실패: {path} — {e}')
            if path == candidates[-1]:
                raise


def write_csv(scenarios, out_path):
    import csv
    with open(str(out_path), 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow([h for h, _ in HEADERS])
        for s in scenarios:
            w.writerow([s['no'], s['testType'], s['scenarioId'], s['screenNo'],
                        s['screenName'], s['category'], s['testName'], s.get('pgmId', ''),
                        s['crudType'], s['preCondition'], s['testData'],
                        s['expectedResult'], s['testResult'], s['remark']])
    print(f'✅ CSV 저장: {out_path}')

# ══════════════════════════════════════════════════════════════════════════════
# 8. 진입점
# ══════════════════════════════════════════════════════════════════════════════

def find_root(start):
    cur = start
    for _ in range(8):
        if (cur / 'src' / 'main' / 'java').exists():
            return cur
        cur = cur.parent
    return start

def main():
    # sys.stdout 인코딩을 UTF-8로 강제 (Tomcat 서비스 계정 환경에서 cp949로 설정될 수 있음)
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    # argv[1] = 프로젝트 루트, argv[2] = 출력 디렉토리, argv[3] = 화면 번호 필터 (쉼표 구분, 빈=전체)
    root    = Path(sys.argv[1]) if len(sys.argv) > 1 else find_root(Path(__file__).parent)
    out_dir_arg = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    source_filter_arg = sys.argv[3].strip() if len(sys.argv) > 3 else ''
    source_filter = set(source_filter_arg.split(',')) if source_filter_arg else None
    if source_filter:
        print(f'소스 필터(argv[3]): {sorted(source_filter)}')

    print(f'프로젝트 루트: {root}')
    if not (root / 'src').exists():
        print('[ERROR] src 디렉토리 없음')
        sys.exit(1)

    # ── 출력 디렉토리 결정 ────────────────────────────────────────────────────
    # 우선순위: argv[2] > copilot_info/ > 프로젝트 루트 > 시스템 임시 폴더
    import tempfile
    out_dir_candidates = [
        out_dir_arg,
        Path(__file__).parent,           # copilot_info/
        root / 'copilot_info',
        root,
        Path(tempfile.gettempdir()),     # C:/Users/.../AppData/Local/Temp (항상 쓰기 가능)
    ]
    out_dir = None
    for candidate in out_dir_candidates:
        if candidate is None:
            continue
        try:
            candidate.mkdir(parents=True, exist_ok=True)
            # 쓰기 권한 테스트
            test_file = candidate / '.write_test'
            test_file.write_text('test', encoding='utf-8')
            test_file.unlink()
            out_dir = candidate
            break
        except Exception:
            continue
    if out_dir is None:
        print('[ERROR] 쓰기 가능한 출력 디렉토리가 없습니다.')
        sys.exit(1)
    print(f'출력 디렉토리: {out_dir}')

    # ── scenario-rules-pur.md 로드 ────────────────────────────────────────────
    md_candidates = [
        Path(__file__).parent / 'scenario-rules-pur.md',
        root / 'copilot_info' / 'scenario-rules-pur.md',
        root / 'scenario-rules-pur.md',
    ]
    md_path = next((p for p in md_candidates if p.exists()), md_candidates[0])
    rules = load_rules_from_md(md_path)
    apply_rules(rules)
    print()

    scenarios = generate_scenarios(root, source_filter)
    if not scenarios:
        print('[ERROR] 시나리오 생성 실패')
        sys.exit(1)

    # ── 양식 파일 탐색 ───────────────────────────────────────────────────────
    template_candidates = [
        root.parent / 'claude' / '양식' / '통합테스트시나리오_양식.xlsx',
        Path(__file__).parent.parent / 'claude' / '양식' / '통합테스트시나리오_양식.xlsx',
        root / 'claude' / '양식' / '통합테스트시나리오_양식.xlsx',
    ]
    template_path = next((p for p in template_candidates if p.exists()), None)
    if template_path:
        print(f'양식 파일 발견: {template_path}')
    else:
        print('양식 파일 없음 → 기본 형식으로 생성')

    today = date.today().strftime('%Y%m%d')
    if HAS_OPENPYXL:
        write_excel(scenarios, out_dir / f'pur_test_scenarios_{today}.xlsx', template_path)
    else:
        write_csv(scenarios, out_dir / f'pur_test_scenarios_{today}.csv')

    from collections import Counter
    cnt = Counter(s['testType'] for s in scenarios)
    print('\n📊 시나리오 통계:')
    for t in ['단위', '통합']:
        if cnt[t]:
            print(f'   {t:10s}: {cnt[t]}건')
    print(f'   {"합계":10s}: {len(scenarios)}건')

if __name__ == '__main__':
    main()
